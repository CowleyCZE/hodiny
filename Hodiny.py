import openpyxl
import datetime
import os
import tkinter as tk
from tkinter import simpledialog, messagebox
from tkcalendar import Calendar

# Cesty k souborům a proměnné
EXCEL_FILE = '/storage/emulated/0/Documents/Hodiny/Hodiny 29-33 Cap.xlsx'
TEMPLATE_SHEET_NAME = 'Týden'

def vypocet_pracovni_doby(cas_zacatku, cas_konce, delka_obeda=1.0):
    format_casu = '%H:%M'
    zacatek = datetime.datetime.strptime(cas_zacatku, format_casu)
    konec = datetime.datetime.strptime(cas_konce, format_casu)
    pracovni_doba = (konec - zacatek).seconds / 3600 - delka_obeda
    return max(pracovni_doba, 0)  # Zajistí, že pracovní doba nebude záporná

def get_or_create_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        if TEMPLATE_SHEET_NAME in wb.sheetnames:
            template_sheet = wb[TEMPLATE_SHEET_NAME]
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = sheet_name
            # Zápis názvu listu do buňky A80
            new_sheet['A80'] = sheet_name
        else:
            raise ValueError(f"Šablona listu '{TEMPLATE_SHEET_NAME}' neexistuje.")
    return wb[sheet_name]

def vybrat_datum():
    def get_date():
        nonlocal selected_date
        selected_date = cal.selection_get()
        top.destroy()

    today = datetime.date.today()
    selected_date = None
    
    top = tk.Toplevel()
    top.title("Vyberte datum")
    cal = Calendar(top, selectmode='day', year=today.year, month=today.month, day=today.day)
    cal.pack(pady=10)
    tk.Button(top, text="Potvrdit", command=get_date).pack()
    
    top.wait_window()
    return selected_date

def zpracovani_dne(wb):
    datum_obj = vybrat_datum()
    if not datum_obj:
        return False

    cislo_tyden = datum_obj.isocalendar()[1]
    sheet_name = f'Týden {cislo_tyden}'
    
    try:
        sheet = get_or_create_sheet(wb, sheet_name)
    except ValueError as e:
        messagebox.showerror("Chyba", str(e))
        return True

    zacatek = simpledialog.askstring("Začátek pracovní doby", "Zadejte čas začátku (HH:MM) nebo 'X' [07:00]:", initialvalue="07:00")
    if zacatek is None:
        return False
    zacatek = zacatek.strip().upper()
    
    if zacatek == 'X':
        konec = 'X'
        pracovni_doba = 'X'
    else:
        try:
            zacatek = zacatek or '07:00'
            datetime.datetime.strptime(zacatek, '%H:%M')
            konec = simpledialog.askstring("Konec pracovní doby", "Zadejte čas konce (HH:MM):")
            if konec is None:
                return False
            datetime.datetime.strptime(konec, '%H:%M')
            delka_obeda = simpledialog.askfloat("Délka oběda", "Zadejte délku oběda v hodinách:")
            if delka_obeda is None:
                return False
            pracovni_doba = vypocet_pracovni_doby(zacatek, konec, delka_obeda)
        except ValueError:
            messagebox.showerror("Chyba", "Neplatný formát času. Použijte HH:MM")
            return True

    index_dne = datum_obj.weekday()
    datum = datum_obj.strftime('%d.%m.%Y')

    sheet.cell(row=7, column=2 + index_dne * 2, value=zacatek)
    sheet.cell(row=7, column=3 + index_dne * 2, value=konec)
    sheet.cell(row=8, column=2 + index_dne * 2, value=pracovni_doba)
    sheet.cell(row=9, column=2 + index_dne * 2, value=pracovni_doba)
    sheet.cell(row=80, column=2 + index_dne * 2, value=datum)

    if zacatek != 'X':
        messagebox.showinfo("Informace", f"Pracovní doba za den {datum} je: {pracovni_doba:.2f} hodin")
    else:
        messagebox.showinfo("Informace", f"Pro den {datum} bylo zadáno 'X'.")

    return True

def main():
    root = tk.Tk()
    root.withdraw()

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        messagebox.showerror("Chyba", f"Soubor '{EXCEL_FILE}' nebyl nalezen.")
        return
    except openpyxl.utils.exceptions.InvalidFileException:
        messagebox.showerror("Chyba", f"Soubor '{EXCEL_FILE}' není platný Excel soubor.")
        return

    while zpracovani_dne(wb):
        pass

    try:
        wb.save(EXCEL_FILE)
        messagebox.showinfo("Informace", "Změny byly uloženy.")
    except PermissionError:
        messagebox.showerror("Chyba", f"Nemáte oprávnění k zápisu do souboru '{EXCEL_FILE}'.")
    except Exception as e:
        messagebox.showerror("Chyba", f"Nastala chyba při ukládání souboru: {str(e)}")

if __name__ == "__main__":
    main()