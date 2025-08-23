"""Hlavní Flask aplikace: routing, inicializace managerů, práce s nastavením.

Tento modul:
 - načítá / ukládá runtime nastavení (pracovní doba, archivovaný týden)
 - instancuje *manager* třídy pro zaměstnance, excel, zálohy a roční souhrn 2025
 - poskytuje web UI (zaměstnanci, záznam času, přehled Excel, zálohy, měsíční report)
 - odesílá aktivní Excel e‑mailem
"""
import json
import smtplib
import os
from datetime import datetime, timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from flask import Flask, flash, g, jsonify, redirect, render_template, request, send_file, session, url_for
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from config import Config
from employee_management import EmployeeManager
from excel_manager import ExcelManager
from hodiny2025_manager import Hodiny2025Manager
from utils.logger import setup_logger
from zalohy_manager import ZalohyManager

logger = setup_logger("app")

app = Flask(__name__)
app.secret_key = Config.SECRET_KEY
Config.init_app(app)


def save_settings_to_file(settings_data):
    """Persistuje nastavení do JSON. Vrací True při úspěchu."""
    try:
        with open(Config.SETTINGS_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump(settings_data, f, indent=4, ensure_ascii=False)
        return True
    except (IOError, TypeError) as e:
        logger.error("Chyba při ukládání nastavení: %s", e, exc_info=True)
        return False


def load_settings_from_file():
    """Načte nastavení z disku nebo vrátí výchozí při chybě / neexistenci."""
    if not Config.SETTINGS_FILE_PATH.exists():
        return Config.get_default_settings()
    try:
        with open(Config.SETTINGS_FILE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        logger.error("Chyba při načítání nastavení: %s", e, exc_info=True)
        return Config.get_default_settings()


@app.before_request
def before_request():
    """Před každým requestem připraví managers + synchronizuje nastavení v session.

    Současně kontroluje, zda nezačal nový týden = případná archivace starého.
    """
    session["settings"] = load_settings_from_file()
    g.employee_manager = EmployeeManager(Config.DATA_PATH)
    g.excel_manager = ExcelManager(Config.EXCEL_BASE_PATH)
    g.zalohy_manager = ZalohyManager(Config.EXCEL_BASE_PATH)
    g.hodiny2025_manager = Hodiny2025Manager(Config.EXCEL_BASE_PATH)

    # Automatická archivace při přechodu na nový týden
    current_week = datetime.now().isocalendar().week
    if g.excel_manager.archive_if_needed(current_week, session["settings"]):
        save_settings_to_file(session["settings"])
        flash(f"Týden {session['settings']['last_archived_week'] - 1} byl archivován.", "info")


@app.teardown_request
def teardown_request(_exception=None):
    """Uzavře případné otevřené workbooky (flush cache)."""
    if hasattr(g, "excel_manager") and g.excel_manager:
        g.excel_manager.close_cached_workbooks()


def _cleanup_temp_files():
    """Vyčistí dočasné soubory starší než 1 hodinu."""
    try:
        import time
        current_time = time.time()
        for temp_file in Config.EXCEL_BASE_PATH.glob("temp_*.xlsx"):
            # Remove temp files older than 1 hour
            if current_time - temp_file.stat().st_mtime > 3600:
                temp_file.unlink()
                logger.info("Odstraněn starý dočasný soubor: %s", temp_file.name)
    except Exception as e:
        logger.error("Chyba při čištění dočasných souborů: %s", e, exc_info=True)


@app.route("/")
def index():
    """Úvodní stránka s rychlými informacemi (aktuální datum + týden)."""
    # Clean up any temporary upload files
    _cleanup_temp_files()
    
    active_filename = Config.EXCEL_TEMPLATE_NAME
    week_num_int = datetime.now().isocalendar().week
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_date_formatted = datetime.now().strftime("%d.%m.%Y")
    
    try:
        excel_exists = g.excel_manager.get_active_file_path().exists()
    except FileNotFoundError:
        excel_exists = False
    
    # Získání projektových informací
    project_name = session.get("settings", {}).get("project_info", {}).get("name", "Nepojmenovaný projekt")
    
    # Získání aktuálního týdenního listu
    current_week_data = None
    try:
        if excel_exists:
            current_week_data = g.excel_manager.get_current_week_data()
    except Exception as e:
        logger.error("Chyba při načítání dat aktuálního týdne: %s", e, exc_info=True)
    
    # Získání výchozích časů pro rychlé zadání
    start_time = session.get("settings", {}).get("start_time", "07:00")
    end_time = session.get("settings", {}).get("end_time", "18:00")
    lunch_duration = session.get("settings", {}).get("lunch_duration", 1.0)
    
    return render_template(
        "index.html",
        active_filename=active_filename,
        week_number=week_num_int,
        current_date=current_date,
        current_date_formatted=current_date_formatted,
        excel_exists=excel_exists,
        project_name=project_name,
        current_week_data=current_week_data,
        start_time=start_time,
        end_time=end_time,
        lunch_duration=lunch_duration,
    )


@app.route("/download")
def download_file():
    """Stáhne aktuálně aktivní Excel soubor."""
    try:
        return send_file(g.excel_manager.get_active_file_path(), as_attachment=True)
    except FileNotFoundError as e:
        logger.error("Chyba při stahování souboru: %s", e, exc_info=True)
        flash("Chyba při stahování souboru.", "error")
        return redirect(url_for("index"))


@app.route("/send_email", methods=["POST"])
def send_email():
    """Odešle aktivní Excel jako přílohu na konfigurovaný e‑mail (SMTP SSL)."""
    try:
        recipient = Config.RECIPIENT_EMAIL or ""
        sender = Config.SMTP_USERNAME or ""
        if not all([recipient, sender, Config.SMTP_PASSWORD, Config.SMTP_SERVER, Config.SMTP_PORT]):
            raise ValueError("SMTP údaje nejsou kompletní.")

        msg = MIMEMultipart()
        msg["Subject"] = f'Výkaz práce - {datetime.now().strftime("%Y-%m-%d")}'
        msg["From"] = sender
        msg["To"] = recipient
        msg.attach(MIMEText("V příloze zasílám výkaz práce.", "plain", "utf-8"))

        with open(g.excel_manager.get_active_file_path(), "rb") as f:
            attachment = MIMEApplication(
                f.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            attachment.add_header(
                "Content-Disposition", "attachment", filename=g.excel_manager.active_filename
            )
            msg.attach(attachment)

        with smtplib.SMTP_SSL(
            Config.SMTP_SERVER, Config.SMTP_PORT, timeout=Config.SMTP_TIMEOUT
        ) as smtp:
            smtp.login(sender, Config.SMTP_PASSWORD if Config.SMTP_PASSWORD is not None else "")
            smtp.send_message(msg)
        flash("Email byl úspěšně odeslán.", "success")
    except (ValueError, smtplib.SMTPException, FileNotFoundError) as e:
        logger.error("Chyba při odesílání emailu: %s", e, exc_info=True)
        flash("Chyba při odesílání emailu.", "error")
    return redirect(url_for("index"))


@app.route("/upload", methods=["GET", "POST"])
def upload_file():
    """Nahrání Excel souborů s kontrolou přepsání existujících souborů."""
    if request.method == "POST":
        # Handle overwrite confirmation
        if "overwrite" in request.form and "filename" in request.form:
            filename = request.form["filename"]
            # This means user confirmed overwrite, but we need the file again
            # Instead, we'll handle this case differently - store file temporarily
            flash("Pro dokončení přepsání prosím znovu vyberte soubor.", "info")
            return redirect(url_for("index"))
        
        # Check if file was uploaded
        if "file" not in request.files:
            flash("Nebyl vybrán žádný soubor.", "error")
            return redirect(url_for("index"))
        
        file = request.files["file"]
        if file.filename == "":
            flash("Nebyl vybrán žádný soubor.", "error")
            return redirect(url_for("index"))
        
        # Validate file extension
        if not file.filename.lower().endswith(".xlsx"):
            flash("Lze nahrávat pouze soubory s příponou .xlsx.", "error")
            return redirect(url_for("index"))
        
        # Secure the filename
        filename = secure_filename(file.filename)
        if not filename:
            flash("Neplatný název souboru.", "error")
            return redirect(url_for("index"))
        
        # Check if file already exists and no overwrite confirmation
        file_path = Config.EXCEL_BASE_PATH / filename
        force_overwrite = request.form.get("force_overwrite") == "true"
        
        if file_path.exists() and not force_overwrite:
            # Store the file temporarily and show confirmation
            try:
                # Validate that it's a valid Excel file by trying to load it
                file.seek(0)
                load_workbook(file)
                file.seek(0)
                
                # Store file temporarily for overwrite confirmation
                temp_path = Config.EXCEL_BASE_PATH / f"temp_{filename}"
                file.save(str(temp_path))
                
                # Pass the temp filename to the confirmation template
                return render_template("upload_confirm.html", 
                                     filename=filename, 
                                     temp_filename=f"temp_{filename}")
                
            except InvalidFileException:
                flash("Soubor není platný Excel soubor (.xlsx).", "error")
                return redirect(url_for("index"))
            except Exception as e:
                logger.error("Chyba při validaci souboru: %s", e, exc_info=True)
                flash("Chyba při nahrávání souboru.", "error")
                return redirect(url_for("index"))
        
        try:
            # Validate that it's a valid Excel file by trying to load it
            file.seek(0)  # Reset file pointer
            load_workbook(file)
            file.seek(0)  # Reset again for saving
            
            # Save the file
            file.save(str(file_path))
            
            flash(f"Soubor '{filename}' byl úspěšně nahrán.", "success")
            logger.info("Nahrán soubor: %s", filename)
            
        except InvalidFileException:
            flash("Soubor není platný Excel soubor (.xlsx).", "error")
        except Exception as e:
            logger.error("Chyba při nahrávání souboru: %s", e, exc_info=True)
            flash("Chyba při nahrávání souboru.", "error")
        
        return redirect(url_for("index"))
    
    # GET request - just redirect to index
    return redirect(url_for("index"))


@app.route("/upload/confirm", methods=["POST"])
def upload_confirm():
    """Potvrzení přepsání existujícího souboru."""
    temp_filename = request.form.get("temp_filename")
    filename = request.form.get("filename")
    
    if not temp_filename or not filename:
        flash("Chyba při zpracování potvrzení.", "error")
        return redirect(url_for("index"))
    
    try:
        temp_path = Config.EXCEL_BASE_PATH / temp_filename
        final_path = Config.EXCEL_BASE_PATH / filename
        
        if not temp_path.exists():
            flash("Dočasný soubor nenalezen. Zkuste nahrání znovu.", "error")
            return redirect(url_for("index"))
        
        # Move temp file to final location
        temp_path.rename(final_path)
        
        flash(f"Soubor '{filename}' byl úspěšně přepsán.", "success")
        logger.info("Přepsán soubor: %s", filename)
        
    except Exception as e:
        logger.error("Chyba při přepisování souboru: %s", e, exc_info=True)
        flash("Chyba při přepisování souboru.", "error")
        
        # Clean up temp file
        try:
            temp_path = Config.EXCEL_BASE_PATH / temp_filename
            if temp_path.exists():
                temp_path.unlink()
        except:
            pass
    
    return redirect(url_for("index"))


@app.route("/zamestnanci", methods=["GET", "POST"])
def manage_employees():
    """Správa zaměstnanců (přidání, výběr pro zapisování, editace, mazání)."""
    if request.method == "POST":
        action = request.form.get("action")
        try:
            if action == "add":
                g.employee_manager.pridat_zamestnance(request.form.get("name", "").strip())
            elif action == "select":
                name = request.form.get("employee_name", "")
                if name in g.employee_manager.vybrani_zamestnanci:
                    g.employee_manager.odebrat_vybraneho_zamestnance(name)
                else:
                    g.employee_manager.pridat_vybraneho_zamestnance(name)
            elif action == "edit":
                old_name = request.form.get("old_name", "").strip()
                new_name = request.form.get("new_name", "").strip()
                g.employee_manager.upravit_zamestnance_podle_jmena(old_name, new_name)
            elif action == "delete":
                g.employee_manager.smazat_zamestnance_podle_jmena(
                    request.form.get("employee_name", "")
                )
        except ValueError as e:
            flash(str(e), "error")
    return render_template("employees.html", employees=g.employee_manager.get_all_employees())


@app.route("/zaznam", methods=["GET", "POST"])
def record_time():
    """Formulář pro zápis pracovní doby / označení volného dne.

    Po úspěšném uložení posune datum na další pracovní den (přeskakuje víkend).
    """
    selected_employees = g.employee_manager.get_vybrani_zamestnanci()
    if not selected_employees:
        flash("Nejsou vybráni žádní zaměstnanci.", "warning")
        return redirect(url_for("manage_employees"))

    current_date = request.args.get("next_date", datetime.now().strftime("%Y-%m-%d"))
    start_time = session["settings"].get("start_time", "07:00")
    end_time = session["settings"].get("end_time", "18:00")
    lunch_duration = str(session["settings"].get("lunch_duration", 1.0))
    is_free_day = False

    if request.method == "POST":
        current_date = request.form.get("date", current_date)
        start_time = request.form.get("start_time", start_time)
        end_time = request.form.get("end_time", end_time)
        lunch_duration = request.form.get("lunch_duration", lunch_duration)
        is_free_day = request.form.get("is_free_day") == "on"

        try:
            date = datetime.strptime(current_date, "%Y-%m-%d").date()
            if is_free_day:
                # Volný den = explicitně 0 hodin, zachová konzistentní záznam
                g.excel_manager.ulozit_pracovni_dobu(
                    current_date, "00:00", "00:00", "0", selected_employees
                )
                g.hodiny2025_manager.zapis_pracovni_doby(
                    current_date, "00:00", "00:00", "0", len(selected_employees)
                )
            else:
                g.excel_manager.ulozit_pracovni_dobu(
                    current_date, start_time, end_time, lunch_duration, selected_employees
                )
                g.hodiny2025_manager.zapis_pracovni_doby(
                    current_date, start_time, end_time, lunch_duration, len(selected_employees)
                )

            flash("Záznam byl úspěšně uložen.", "success")
            next_day = date + timedelta(days=1)
            while next_day.weekday() >= 5:
                next_day += timedelta(days=1)
            return redirect(url_for("record_time", next_date=next_day.strftime("%Y-%m-%d")))
        except (ValueError, IOError, FileNotFoundError) as e:
            flash(str(e), "error")

    return render_template(
        "record_time.html",
        selected_employees=selected_employees,
        current_date=current_date,
        start_time=start_time,
        end_time=end_time,
        lunch_duration=lunch_duration,
        is_free_day=is_free_day,
    )


@app.route("/excel_viewer", methods=["GET"])
def excel_viewer():
    """Read‑only náhled omezeného počtu řádků aktivního Excelu (výkon)."""
    requested_file = request.args.get("file")
    active_sheet_name = request.args.get("sheet")
    data, sheet_names = [], []
    base_path = Config.EXCEL_BASE_PATH
    excel_files = sorted([p.name for p in base_path.glob("*.xlsx")])
    if not excel_files:
        flash("Nenalezeny žádné Excel soubory.", "warning")
        return render_template(
            "excel_viewer.html",
            excel_files=[],
            selected_file=None,
            sheet_names=[],
            active_sheet=None,
            data=[],
        )

    selected_file = (
        requested_file if requested_file in excel_files else g.excel_manager.active_filename
    )
    selected_path = base_path / selected_file
    try:
        wb = load_workbook(selected_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        if not sheet_names:
            flash("Vybraný soubor nemá žádné listy.", "warning")
            return render_template(
                "excel_viewer.html",
                excel_files=excel_files,
                selected_file=selected_file,
                sheet_names=[],
                active_sheet=None,
                data=[],
            )
        active_sheet_name = (
            active_sheet_name if active_sheet_name in sheet_names else sheet_names[0]
        )
        sheet = wb[active_sheet_name]
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= Config.MAX_ROWS_TO_DISPLAY_EXCEL_VIEWER:
                break
            data.append([str(c) if c is not None else "" for c in row])
        wb.close()
    except (FileNotFoundError, InvalidFileException) as e:
        flash(f"Chyba při zobrazení souboru: {e}", "error")
        return redirect(url_for("index"))

    return render_template(
        "excel_viewer.html",
        excel_files=excel_files,
        selected_file=selected_file,
        sheet_names=sheet_names,
        active_sheet=active_sheet_name,
        data=data,
    )


@app.route("/excel_editor", methods=["GET", "POST"])
def excel_editor():
    """Editable náhled Excel souborů s možností úprav přímo v prohlížeči."""
    if request.method == "POST":
        # Handle cell edit
        try:
            file_name = request.form.get("file")
            sheet_name = request.form.get("sheet")
            row = int(request.form.get("row"))
            col = int(request.form.get("col"))
            value = request.form.get("value")
            
            if not file_name or not sheet_name:
                flash("Chybí název souboru nebo listu.", "error")
                return redirect(url_for("excel_editor"))
            
            base_path = Config.EXCEL_BASE_PATH
            file_path = base_path / file_name
            
            # Load workbook for editing
            wb = load_workbook(file_path)
            sheet = wb[sheet_name]
            
            # Update cell value
            sheet.cell(row=row, column=col, value=value)
            
            # Save the workbook
            wb.save(file_path)
            wb.close()
            
            flash("Buňka byla úspěšně aktualizována.", "success")
            return redirect(url_for("excel_editor", file=file_name, sheet=sheet_name))
            
        except Exception as e:
            flash(f"Chyba při ukládání: {e}", "error")
            return redirect(url_for("excel_editor"))
    
    # GET request - display the editor
    requested_file = request.args.get("file")
    active_sheet_name = request.args.get("sheet")
    data, sheet_names = [], []
    base_path = Config.EXCEL_BASE_PATH
    excel_files = sorted([p.name for p in base_path.glob("*.xlsx")])
    
    if not excel_files:
        flash("Nenalezeny žádné Excel soubory.", "warning")
        return render_template(
            "excel_editor.html",
            excel_files=[],
            selected_file=None,
            sheet_names=[],
            active_sheet=None,
            data=[],
        )

    selected_file = (
        requested_file if requested_file in excel_files else g.excel_manager.active_filename
    )
    selected_path = base_path / selected_file
    
    try:
        wb = load_workbook(selected_path, data_only=False)
        sheet_names = wb.sheetnames
        if not sheet_names:
            flash("Vybraný soubor nemá žádné listy.", "warning")
            return render_template(
                "excel_editor.html",
                excel_files=excel_files,
                selected_file=selected_file,
                sheet_names=[],
                active_sheet=None,
                data=[],
            )
        
        active_sheet_name = (
            active_sheet_name if active_sheet_name in sheet_names else sheet_names[0]
        )
        sheet = wb[active_sheet_name]
        
        # Prepare data with row/column indices for editing
        data_with_coords = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_idx > Config.MAX_ROWS_TO_DISPLAY_EXCEL_VIEWER:
                break
            row_data = []
            for col_idx, cell_value in enumerate(row, 1):
                row_data.append({
                    'value': str(cell_value) if cell_value is not None else "",
                    'row': row_idx,
                    'col': col_idx
                })
            data_with_coords.append(row_data)
        
        wb.close()
        
    except (FileNotFoundError, InvalidFileException) as e:
        flash(f"Chyba při zobrazení souboru: {e}", "error")
        return redirect(url_for("index"))

    return render_template(
        "excel_editor.html",
        excel_files=excel_files,
        selected_file=selected_file,
        sheet_names=sheet_names,
        active_sheet=active_sheet_name,
        data=data_with_coords,
    )


@app.route("/settings", methods=["GET", "POST"])
def settings_page():
    """Nastavení výchozí pracovní doby (start/end + oběd)."""
    if request.method == "POST":
        try:
            settings_to_save = session["settings"].copy()
            settings_to_save.update(
                {
                    "start_time": request.form["start_time"],
                    "end_time": request.form["end_time"],
                    "lunch_duration": float(request.form["lunch_duration"].replace(",", ".")),
                }
            )
            if not save_settings_to_file(settings_to_save):
                raise IOError("Nepodařilo se uložit nastavení.")
            session["settings"] = settings_to_save
            flash("Nastavení bylo úspěšně uloženo.", "success")
        except (ValueError, IOError) as e:
            flash(str(e), "error")

    return render_template("settings_page.html", settings=session.get("settings", {}))


@app.route("/zalohy", methods=["GET", "POST"])
def zalohy():
    """Správa záloh (půjček / plateb) pro zaměstnance."""
    if request.method == "POST":
        try:
            form = request.form
            amount = float(form["amount"].replace(",", "."))
            g.zalohy_manager.add_or_update_employee_advance(
                form["employee_name"],
                amount,
                form["currency"],
                form["option"],
                form["date"],
            )
            flash("Záloha byla úspěšně uložena.", "success")
        except (ValueError, IOError) as e:
            flash(str(e), "error")

    return render_template(
        "zalohy.html",
        employees=g.employee_manager.zamestnanci,
        options=g.zalohy_manager.get_option_names(),
        current_date=datetime.now().strftime("%Y-%m-%d"),
    )


@app.route("/monthly_report", methods=["GET", "POST"])
def monthly_report_route():
    """Generuje měsíční agregace z týdenních listů podle zvolených zaměstnanců."""
    report_data = None
    if request.method == "POST":
        try:
            month = int(request.form["month"])
            year = int(request.form["year"])
            employees = request.form.getlist("employees") or None
            report_data = g.excel_manager.generate_monthly_report(
                month, year, employees
            )
            if not report_data:
                flash("Nebyly nalezeny žádné záznamy.", "info")
        except (ValueError, FileNotFoundError) as e:
            flash(str(e), "error")

    employee_names = [emp["name"] for emp in g.employee_manager.get_all_employees()]
    return render_template(
        "monthly_report.html",
        employee_names=employee_names,
        report_data=report_data,
        current_month=datetime.now().month,
        current_year=datetime.now().year,
    )


def load_dynamic_config():
    """Načte dynamickou konfiguraci pro ukládání do XLSX souborů."""
    if not Config.CONFIG_FILE_PATH.exists():
        return {}
    try:
        with open(Config.CONFIG_FILE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError) as e:
        logger.error("Chyba při načítání dynamické konfigurace: %s", e, exc_info=True)
        return {}


def save_dynamic_config(config_data):
    """Uloží dynamickou konfiguraci do JSON. Vrací True při úspěchu."""
    try:
        with open(Config.CONFIG_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump(config_data, f, indent=4, ensure_ascii=False)
        return True
    except (IOError, TypeError) as e:
        logger.error("Chyba při ukládání dynamické konfigurace: %s", e, exc_info=True)
        return False


# API endpoints pro dynamickou konfiguraci
@app.route("/api/settings", methods=["GET"])
def api_get_settings():
    """Vrací aktuální obsah souboru config.json."""
    try:
        config = load_dynamic_config()
        return jsonify(config)
    except Exception as e:
        logger.error("Chyba při načítání API nastavení: %s", e, exc_info=True)
        return jsonify({"error": "Chyba při načítání nastavení"}), 500


@app.route("/api/settings", methods=["POST"])
def api_save_settings():
    """Přijímá nová nastavení ve formátu JSON a ukládá je do config.json."""
    try:
        config_data = request.get_json()
        if not config_data:
            return jsonify({"error": "Žádná data nebyla odeslána"}), 400
        
        if not save_dynamic_config(config_data):
            return jsonify({"error": "Nepodařilo se uložit nastavení"}), 500
        
        return jsonify({"success": True, "message": "Nastavení bylo úspěšně uloženo"})
    except Exception as e:
        logger.error("Chyba při ukládání API nastavení: %s", e, exc_info=True)
        return jsonify({"error": "Chyba při ukládání nastavení"}), 500


@app.route("/api/files", methods=["GET"])
def api_get_files():
    """Prohledá Excel složku a vrátí seznam všech .xlsx souborů."""
    try:
        excel_files = sorted([p.name for p in Config.EXCEL_BASE_PATH.glob("*.xlsx")])
        return jsonify({"files": excel_files})
    except Exception as e:
        logger.error("Chyba při načítání Excel souborů: %s", e, exc_info=True)
        return jsonify({"error": "Chyba při načítání souborů"}), 500


@app.route("/api/sheets/<filename>", methods=["GET"])
def api_get_sheets(filename):
    """Přijme název souboru a vrátí seznam názvů všech listů."""
    try:
        file_path = Config.EXCEL_BASE_PATH / filename
        if not file_path.exists():
            return jsonify({"error": "Soubor nenalezen"}), 404
        
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        return jsonify({"sheets": sheet_names})
    except Exception as e:
        logger.error("Chyba při načítání listů ze souboru %s: %s", filename, e, exc_info=True)
        return jsonify({"error": "Chyba při načítání listů"}), 500


@app.route("/api/sheet_content/<filename>/<sheetname>", methods=["GET"])
def api_get_sheet_content(filename, sheetname):
    """Vrátí obsah zadaného listu ve formátu JSON (pole polí)."""
    try:
        file_path = Config.EXCEL_BASE_PATH / filename
        if not file_path.exists():
            return jsonify({"error": "Soubor nenalezen"}), 404
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if sheetname not in wb.sheetnames:
            wb.close()
            return jsonify({"error": "List nenalezen"}), 404
        
        sheet = wb[sheetname]
        data = []
        
        # Omezíme počet řádků pro výkon
        max_rows = min(sheet.max_row, Config.MAX_ROWS_TO_DISPLAY_EXCEL_VIEWER)
        max_cols = min(sheet.max_column, 26)  # Omezení na sloupce A-Z
        
        for row_idx in range(1, max_rows + 1):
            row_data = []
            for col_idx in range(1, max_cols + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            data.append(row_data)
        
        wb.close()
        
        # Vrátíme také informace o rozměrech pro frontend
        return jsonify({
            "data": data,
            "rows": max_rows,
            "cols": max_cols
        })
    except Exception as e:
        logger.error("Chyba při načítání obsahu listu %s/%s: %s", filename, sheetname, e, exc_info=True)
        return jsonify({"error": "Chyba při načítání obsahu listu"}), 500


@app.route("/nastaveni")
def nastaveni_page():
    """Stránka pro dynamické nastavení ukládání dat do XLSX souborů."""
    return render_template("nastaveni.html")


@app.route("/api/files/rename", methods=["POST"])
def rename_file():
    """API endpoint pro přejmenování XLSX souborů."""
    try:
        data = request.get_json()
        old_filename = data.get('old_filename')
        new_filename = data.get('new_filename')
        
        if not old_filename or not new_filename:
            return jsonify({"success": False, "error": "Chybí název souboru"}), 400
            
        # Kontrola, že jde o xlsx soubory
        if not old_filename.endswith('.xlsx') or not new_filename.endswith('.xlsx'):
            return jsonify({"success": False, "error": "Pouze .xlsx soubory mohou být přejmenovány"}), 400
            
        old_path = Config.EXCEL_BASE_PATH / old_filename
        new_path = Config.EXCEL_BASE_PATH / new_filename
        
        # Kontrola existence starého souboru
        if not old_path.exists():
            return jsonify({"success": False, "error": f"Soubor {old_filename} neexistuje"}), 404
            
        # Kontrola, že nový soubor neexistuje
        if new_path.exists():
            return jsonify({"success": False, "error": f"Soubor {new_filename} již existuje"}), 409
            
        # Přejmenování souboru
        old_path.rename(new_path)
        
        logger.info("Soubor %s byl přejmenován na %s", old_filename, new_filename)
        return jsonify({"success": True, "message": f"Soubor byl úspěšně přejmenován na {new_filename}"})
        
    except Exception as e:
        logger.error("Chyba při přejmenování souboru: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/quick_time_entry", methods=["POST"])
def quick_time_entry():
    """API endpoint pro rychlé zadání pracovní doby z hlavní stránky."""
    try:
        data = request.get_json()
        date = data.get('date')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        lunch_duration = data.get('lunch_duration', '1.0')
        is_free_day = data.get('is_free_day', False)
        
        if not date:
            return jsonify({"success": False, "error": "Chybí datum"}), 400
            
        selected_employees = g.employee_manager.get_vybrani_zamestnanci()
        if not selected_employees:
            return jsonify({"success": False, "error": "Nejsou vybráni žádní zaměstnanci"}), 400
        
        try:
            # Validace data
            datetime.strptime(date, "%Y-%m-%d")
            
            if is_free_day:
                # Volný den
                g.excel_manager.ulozit_pracovni_dobu(
                    date, "00:00", "00:00", "0", selected_employees
                )
                g.hodiny2025_manager.zapis_pracovni_doby(
                    date, "00:00", "00:00", "0", len(selected_employees)
                )
                message = f"Volný den pro {date} byl zaznamenán pro {len(selected_employees)} zaměstnanců"
            else:
                if not start_time or not end_time:
                    return jsonify({"success": False, "error": "Chybí čas začátku nebo konce"}), 400
                    
                g.excel_manager.ulozit_pracovni_dobu(
                    date, start_time, end_time, lunch_duration, selected_employees
                )
                g.hodiny2025_manager.zapis_pracovni_doby(
                    date, start_time, end_time, lunch_duration, len(selected_employees)
                )
                message = f"Pracovní doba pro {date} byla zaznamenána pro {len(selected_employees)} zaměstnanců"
            
            return jsonify({"success": True, "message": message})
            
        except ValueError as e:
            return jsonify({"success": False, "error": f"Neplatné datum nebo čas: {e}"}), 400
            
    except Exception as e:
        logger.error("Chyba při rychlém zadání času: %s", e, exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
