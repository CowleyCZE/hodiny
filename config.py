# config.py
import os
import secrets
from pathlib import Path
import logging
from openpyxl import Workbook

class Config:
    IS_PYTHONANYWHERE = "PYTHONANYWHERE_SITE" in os.environ
    SECRET_KEY = os.environ.get("SECRET_KEY", secrets.token_hex(32))

    BASE_DIR = Path(os.path.expanduser("~/hodiny") if IS_PYTHONANYWHERE else os.environ.get("HODINY_BASE_DIR", ".")).resolve()
    DATA_PATH = Path(os.environ.get("HODINY_DATA_PATH", BASE_DIR / "data"))
    EXCEL_BASE_PATH = Path(os.environ.get("HODINY_EXCEL_PATH", BASE_DIR / "excel"))
    EXCEL_TEMPLATE_NAME = "Hodiny_Cap.xlsx"
    SETTINGS_FILE_PATH = DATA_PATH / "settings.json"

    SMTP_SERVER = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
    SMTP_PORT = int(os.environ.get("SMTP_PORT", 465))
    SMTP_USERNAME = os.environ.get("SMTP_USERNAME")
    SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD")
    RECIPIENT_EMAIL = os.environ.get("RECIPIENT_EMAIL")

    DEFAULT_APP_NAME = 'Evidence pracovní doby'
    SMTP_TIMEOUT = 60
    MAX_ROWS_TO_DISPLAY_EXCEL_VIEWER = 500
    EXCEL_EMPLOYEE_START_ROW = 9
    EXCEL_WEEK_SHEET_TEMPLATE_NAME = "Týden"
    EXCEL_ADVANCES_SHEET_NAME = "Zálohy"
    DEFAULT_ADVANCE_OPTION_1 = "Option 1"
    DEFAULT_ADVANCE_OPTION_2 = "Option 2"
    DEFAULT_ADVANCE_OPTION_3 = "Option 3"
    DEFAULT_ADVANCE_OPTION_4 = "Option 4"
    EMPLOYEE_NAME_VALIDATION_REGEX = r"^[\w\s\-.ěščřžýáíéúůďťňĚŠČŘŽÝÁÍÉÚŮĎŤŇ]+$"
    EMPLOYEE_NAME_MAX_LENGTH = 100

    @dataclass
    class TimeConfig:
        start_time: str = "07:00"
        end_time: str = "18:00"
        lunch_duration: float = 1.0

    DEFAULT_TIME_CONFIG = TimeConfig()

    @classmethod
    def get_default_settings(cls):
        return {
            "start_time": cls.DEFAULT_TIME_CONFIG.start_time,
            "end_time": cls.DEFAULT_TIME_CONFIG.end_time,
            "lunch_duration": cls.DEFAULT_TIME_CONFIG.lunch_duration,
            "project_info": {"name": "", "start_date": "", "end_date": ""},
            "last_archived_week": 0,  # Sledování posledního archivovaného týdne
        }

    @classmethod
    def init_app(cls, app):
        cls.DATA_PATH.mkdir(parents=True, exist_ok=True)
        cls.EXCEL_BASE_PATH.mkdir(parents=True, exist_ok=True)

        template_path = cls.EXCEL_BASE_PATH / cls.EXCEL_TEMPLATE_NAME
        if not template_path.exists():
            try:
                wb = Workbook()
                wb.create_sheet(cls.EXCEL_WEEK_SHEET_TEMPLATE_NAME)
                zalohy_sheet = wb.create_sheet(cls.EXCEL_ADVANCES_SHEET_NAME)
                zalohy_sheet["B80"] = cls.DEFAULT_ADVANCE_OPTION_1
                zalohy_sheet["D80"] = cls.DEFAULT_ADVANCE_OPTION_2
                zalohy_sheet["F80"] = cls.DEFAULT_ADVANCE_OPTION_3
                zalohy_sheet["H80"] = cls.DEFAULT_ADVANCE_OPTION_4
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
                wb.save(template_path)
            except Exception as e:
                logging.error(f"Nepodařilo se vytvořit šablonu {template_path}: {e}", exc_info=True)

        log_dir = cls.BASE_DIR / "logs"
        log_dir.mkdir(exist_ok=True)

        app.config.update(
            SECRET_KEY=cls.SECRET_KEY,
            UPLOAD_FOLDER=str(cls.EXCEL_BASE_PATH),
            SESSION_COOKIE_SECURE=not app.debug,
            SESSION_COOKIE_HTTPONLY=True,
            PERMANENT_SESSION_LIFETIME=3600,
            PROPAGATE_EXCEPTIONS=cls.IS_PYTHONANYWHERE,
            PREFERRED_URL_SCHEME="https" if cls.IS_PYTHONANYWHERE else "http"
        )
        return app
