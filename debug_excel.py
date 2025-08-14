#!/usr/bin/env python3
"""
Debug script pro inspekci Excel souboru Hodiny2025.xlsx
"""
from openpyxl import load_workbook

def debug_excel_file():
    file_path = "/workspaces/hodiny/excel/Hodiny2025.xlsx"
    
    try:
        workbook = load_workbook(file_path)
        print(f"🔍 DEBUGGING EXCEL SOUBORU: {file_path}")
        print("=" * 60)
        
        # Zobrazit všechny listy
        print(f"📋 Celkem listů: {len(workbook.sheetnames)}")
        print(f"📋 Názvy listů: {workbook.sheetnames[:10]}...")  # Prvních 10
        
        # Zkontroluj list 08hod25
        if "08hod25" in workbook.sheetnames:
            sheet = workbook["08hod25"]
            print(f"\n📊 ANALÝZA LISTU: 08hod25")
            print("-" * 40)
            
            # Zkontroluj řádek 15 (13. srpen 2025)
            row = 15
            print(f"🎯 Kontrola řádku {row} (13. srpen):")
            
            for col in range(1, 15):  # A až N
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                if value is not None:
                    print(f"   {chr(64+col)}{row}: {value} (typ: {type(value).__name__})")
            
            # Zkontroluj řádek 16 (14. srpen 2025)
            row = 16
            print(f"🎯 Kontrola řádku {row} (14. srpen):")
            
            for col in range(1, 15):  # A až N
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                if value is not None:
                    print(f"   {chr(64+col)}{row}: {value} (typ: {type(value).__name__})")
                    
            # Zkontroluj konkrétní buňky
            print(f"\n🔍 KONKRÉTNÍ HODNOTY:")
            print(f"   E15 (začátek): {sheet['E15'].value}")
            print(f"   F15 (oběd): {sheet['F15'].value}")
            print(f"   G15 (konec): {sheet['G15'].value}")
            print(f"   H15 (celkem): {sheet['H15'].value}")
            print(f"   I15 (přesčasy): {sheet['I15'].value}")
            print(f"   M15 (zaměstnanci): {sheet['M15'].value}")
            print(f"   N15 (celkem_všichni): {sheet['N15'].value}")
            
    except Exception as e:
        print(f"❌ Chyba při načítání souboru: {e}")

if __name__ == "__main__":
    debug_excel_file()
