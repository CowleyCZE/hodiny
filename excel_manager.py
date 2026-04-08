"""Správa týdenního Excel souboru (šablona + dynamické listy týdnů, reporty).

Zodpovědnosti:
 - Lazy/cached otevření workbooku (thread‑safe)
 - Archivace starého týdne a vyčištění listů
 - Zápis denních hodin pro vybrané zaměstnance do listu konkrétního týdne
 - Generace měsíčního reportu agregací z týdenních listů
"""
import logging
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from threading import RLock

from openpyxl import load_workbook
from config import Config
from services.excel_config_service import get_configured_cells
from services.excel_metadata_service import load_metadata, save_metadata, set_file_category
from services.excel_report_service import generate_monthly_report_from_workbook
from services.excel_week_service import (
    archive_active_week_file,
    create_week_sheet_from_template,
    ensure_week_sheet,
    get_current_week_preview,
    get_or_create_weekly_file,
    open_weekly_workbook,
    write_time_entry_to_sheet,
)

try:
    from utils.logger import setup_logger

    logger = setup_logger("excel_manager")
except ImportError:
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger("excel_manager")


class ExcelManager:
    """Správa a manipulace s hlavním Excel souborem."""

    def __init__(self, base_path, hodiny2025_manager=None):
        """base_path: adresář kde je hlavní soubor (Hodiny_Cap.xlsx)."""
        self.base_path = Path(base_path)
        self.active_filename = Config.EXCEL_TEMPLATE_NAME
        self.active_file_path = self.base_path / self.active_filename
        self.current_project_name = None
        self._file_lock = RLock()
        self._workbook_cache = {}
        self._metadata_path = self.base_path / "metadata.json"
        self.hodiny2025_manager = hodiny2025_manager
        logger.info("ExcelManager inicializován pro: %s", self.active_file_path)

    def get_active_file_path(self):
        """Cesta k aktivnímu Excel souboru (šablona)."""
        return self.active_file_path

    def file_exists(self):
        """Ověří, zda hlavní Excel soubor (šablona) existuje na disku."""
        return self.active_file_path.exists()

    def get_active_filename(self):
        """Vrátí název aktivního souboru."""
        return self.active_filename

    def _load_dynamic_config(self):
        """Načte dynamickou konfiguraci pro ukládání do XLSX souborů."""
        from services.excel_config_service import load_dynamic_excel_config

        return load_dynamic_excel_config()

    def _get_cell_coordinates(self, field_key, sheet_name=None, data_type="weekly_time"):
        """Vrátí seznam (row, col) souřadnic pro daný field z dynamické konfigurace.

        Args:
            field_key: Klíč pole z konfigurace (např. 'start_time', 'date')
            sheet_name: Název listu, pokud chceme ověřit shodu
            data_type: Typ dat ('weekly_time', 'advances', 'monthly_time', 'projects')

        Returns:
            list: Seznam (row, col) souřadnic nebo prázdný seznam pokud není nakonfigurováno
        """
        return get_configured_cells(data_type, field_key, self.active_filename, sheet_name=sheet_name)

    @contextmanager
    def _get_workbook(self, filename=None, read_only=False):
        """Context manager vracející workbook (cache pro read_write režim).

        filename: název souboru v base_path (pokud None, použije se aktivní),
        read_only=True -> vždy otevře nový objekt (neukládá do cache),
        jinak recykluje instanci pro snížení IO.
        """
        target_filename = filename or self.active_filename
        target_path = self.base_path / target_filename
        cache_key = str(target_path.absolute())

        wb = None
        with self._file_lock:
            if cache_key in self._workbook_cache:
                try:
                    wb = self._workbook_cache[cache_key]
                    if wb:
                        _ = wb.sheetnames  # Test if workbook is alive
                except Exception:
                    wb = None
            if wb is None:
                if not target_path.exists():
                    raise FileNotFoundError(f"Soubor '{target_filename}' nenalezen.")
                try:
                    wb = load_workbook(
                        filename=str(target_path),
                        read_only=read_only,
                        data_only=not read_only,
                    )
                    if not read_only:
                        self._workbook_cache[cache_key] = wb
                except Exception as e:
                    raise IOError(f"Chyba při otevírání souboru '{target_filename}': {e}") from e
        try:
            yield wb
        finally:
            if read_only and wb:
                wb.close()

    def update_cell(self, filename, sheet_name, row, col, value):
        """Bezpečně aktualizuje hodnotu jedné buňky v zadaném souboru a listu."""
        try:
            with self._file_lock:
                with self._get_workbook(filename=filename, read_only=False) as wb:
                    if sheet_name not in wb.sheetnames:
                        raise ValueError(f"List '{sheet_name}' v souboru '{filename}' neexistuje.")

                    sheet = wb[sheet_name]

                    # Sanitize value to prevent formula injection
                    if value and isinstance(value, str) and value.strip().startswith(("=", "+", "-", "@")):
                        value = "'" + value

                    sheet.cell(row=row, column=col, value=value)
                    # Workbook zůstává v cache a uloží se při close_cached_workbooks.
            logger.info("Buňka [%d, %d] v %s/%s aktualizována na: %s", row, col, filename, sheet_name, value)
            return True
        except Exception as e:
            logger.error("Chyba při aktualizaci buňky v %s: %s", filename, e, exc_info=True)
            return False

    def archive_if_needed(self, current_week_number, settings):
        """Archivuje minulé týdny při posunu čísla týdne."""
        last_archived_week = settings.get("last_archived_week", 0)
        if current_week_number <= last_archived_week:
            return False

        try:
            with self._get_workbook() as wb:
                if not wb:
                    raise IOError("Workbook se nepodařilo otevřít pro archivaci.")
                archive_active_week_file(self.active_file_path, wb, current_week_number, last_archived_week)

            settings["last_archived_week"] = current_week_number
            return True
        except (IOError, FileNotFoundError) as e:
            logger.error("Chyba při archivaci souboru: %s", e, exc_info=True)
            return False

    def ulozit_pracovni_dobu(self, date_str, start_time_str, end_time_str, lunch_duration_str, employees):
        """Zapíše pracovní dobu do týdenního souboru i do aktivního workbooku aplikace."""
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            week_number = date_obj.isocalendar().week

            if not employees:
                raise ValueError("Seznam zaměstnanců nesmí být prázdný.")

            # Zachovej historické týdenní soubory kvůli kompatibilitě a testům.
            weekly_file_path = self._get_or_create_weekly_file(week_number)
            weekly_workbook = self._get_weekly_workbook(weekly_file_path)
            if not weekly_workbook:
                raise IOError(f"Workbook {weekly_file_path} se nepodařilo otevřít.")

            try:
                weekly_sheet_name, weekly_sheet = ensure_week_sheet(weekly_workbook, week_number)
                write_time_entry_to_sheet(
                    weekly_sheet,
                    weekly_sheet_name,
                    date_obj,
                    start_time_str,
                    end_time_str,
                    lunch_duration_str,
                    employees,
                    self._get_cell_coordinates,
                    self.current_project_name,
                )
                weekly_workbook.save(weekly_file_path)
            finally:
                weekly_workbook.close()

            # Aktivní soubor drží současné UI a reportovací flow.
            with self._get_workbook(filename=self.active_filename, read_only=False) as workbook:
                if not workbook:
                    raise IOError(f"Workbook {self.active_file_path} se nepodařilo otevřít.")

                sheet_name = f"{Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME} {week_number}"
                if sheet_name not in workbook.sheetnames:
                    self._create_week_sheet_from_template(workbook, sheet_name)

                sheet = workbook[sheet_name]
                write_time_entry_to_sheet(
                    sheet,
                    sheet_name,
                    date_obj,
                    start_time_str,
                    end_time_str,
                    lunch_duration_str,
                    employees,
                    self._get_cell_coordinates,
                    self.current_project_name,
                )

            if self.hodiny2025_manager:
                try:
                    self.hodiny2025_manager.zapis_pracovni_doby(
                        date_str, start_time_str, end_time_str, lunch_duration_str, len(employees)
                    )
                    logger.info("Data synchronizována s Hodiny2025Manager pro %s.", date_str)
                except Exception as sync_err:
                    logger.error("Chyba při synchronizaci s Hodiny2025Manager: %s", sync_err)

            logger.info("Uložena pracovní doba pro %s do %s.", date_str, weekly_file_path.name)
            return True
        except (FileNotFoundError, ValueError, IOError) as e:
            logger.error("Chyba při ukládání pracovní doby: %s", e, exc_info=True)
            return False
        except Exception as e:
            logger.error("Neočekávaná chyba při ukládání pracovní doby: %s", e, exc_info=True)
            return False

    def _get_or_create_weekly_file(self, week_number):
        """Získá cestu k týdennímu souboru, vytvoří ho pokud neexistuje."""
        return get_or_create_weekly_file(self.base_path, self.active_file_path, week_number)

    def _find_previous_weekly_file(self, current_week):
        """Najde nejnovější týdenní soubor před aktuálním týdnem."""
        from services.excel_week_service import find_previous_weekly_file

        return find_previous_weekly_file(self.base_path, self.active_file_path.stem, current_week)

    def _get_weekly_workbook(self, weekly_file_path):
        """Otevře týdenní workbook pro práci."""
        return open_weekly_workbook(weekly_file_path)

    def _create_week_sheet_from_template(self, workbook, sheet_name):
        """Vytvoří nový list zkopírováním šablony 'Týden' v rámci workbooku."""
        create_week_sheet_from_template(workbook, sheet_name)

    def _zapsat_data_do_listu(
        self, sheet, sheet_name, date_obj, start_time_str, end_time_str, lunch_duration_str, employees
    ):
        """Pomocná metoda pro zápis dat do konkrétního listu."""
        write_time_entry_to_sheet(
            sheet,
            sheet_name,
            date_obj,
            start_time_str,
            end_time_str,
            lunch_duration_str,
            employees,
            self._get_cell_coordinates,
            self.current_project_name,
        )

    def close_cached_workbooks(self):
        """Flush + zavření všech workbooků v cache (volat při ukončení requestu)."""
        with self._file_lock:
            for path_str, wb in self._workbook_cache.items():
                try:
                    if wb:
                        wb.save(path_str)
                        wb.close()
                except Exception as e:
                    logger.error(
                        "Chyba při ukládání/zavírání workbooku %s: %s",
                        path_str,
                        e,
                        exc_info=True,
                    )
            self._workbook_cache.clear()

    def ziskej_cislo_tydne(self, datum):
        """Vrátí ISO kalendář (year, week, weekday) nebo None při chybě."""
        try:
            datum_obj = datetime.strptime(datum, "%Y-%m-%d") if isinstance(datum, str) else datum
            return datum_obj.isocalendar()
        except (ValueError, TypeError) as e:
            logger.error("Chyba při zpracování data '%s': %s", datum, e)
            return None

    def generate_monthly_report(self, month, year, employees=None):
        """Agreguje hodiny / volné dny z týdenních listů pro daný měsíc."""
        if not (1 <= month <= 12 and 2000 <= year <= 2100):
            raise ValueError("Neplatný měsíc nebo rok.")

        report_data = {}
        try:
            with self._get_workbook(read_only=True) as workbook:
                if not workbook:
                    raise IOError("Workbook se nepodařilo otevřít pro report.")
                report_data = generate_monthly_report_from_workbook(workbook, month, year, employees)
        except (FileNotFoundError, IOError) as e:
            logger.error("Chyba při generování měsíčního reportu: %s", e, exc_info=True)
            return {}
        return report_data

    def _get_monthly_sheets(self, workbook, month, year):
        """Generátor pro listy, které spadají do daného měsíce a roku."""
        from services.excel_report_service import get_monthly_sheets

        yield from get_monthly_sheets(workbook, month, year)

    def _process_sheet_for_report(self, sheet, employees, report_data, month, year):
        """Zpracuje jeden list a agreguje data do report_data."""
        from services.excel_report_service import process_sheet_for_report

        process_sheet_for_report(sheet, employees, report_data, month, year)

    def update_project_info(self, project_name, _start_date, _end_date):
        """Aktualizuje runtime informace o projektu pro následný zápis do listů."""
        self.current_project_name = project_name or None
        return True

    def get_current_week_data(self, week_number=None):
        """Získá data z hlavního souboru (Hodiny_Cap.xlsx)."""
        try:
            current_week = week_number or datetime.now().isocalendar().week

            with self._get_workbook(read_only=True) as wb:
                if not wb:
                    return None

                return get_current_week_preview(wb, current_week)

        except Exception as e:
            logger.error("Chyba při načítání dat aktuálního týdne: %s", e, exc_info=True)
            return None

    def _load_metadata(self):
        """Načte metadata souborů z JSON souboru."""
        with self._file_lock:
            return load_metadata(self._metadata_path)

    def _save_metadata(self, metadata):
        """Uloží metadata souborů do JSON souboru."""
        with self._file_lock:
            save_metadata(self._metadata_path, metadata)

    def get_all_metadata(self):
        """Vrátí všechna metadata."""
        return self._load_metadata()

    def set_category(self, filename, category):
        """Nastaví kategorii pro daný soubor."""
        with self._file_lock:
            set_file_category(self._metadata_path, filename, category)
