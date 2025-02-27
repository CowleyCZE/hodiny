import os
from openpyxl import load_workbook, Workbook
from datetime import datetime
import calendar

class ExcelManager2024:
    def __init__(self, base_path):
        self.base_path = base_path
        self.file_path = os.path.join(self.base_path, 'Hodiny2024.xlsx')  # Opravena cesta k souboru

    def ulozit_pracovni_dobu(self, date, start_time, end_time, lunch_duration):
        workbook = self._load_or_create_workbook()
        sheet_name = self._get_sheet_name(date)
        sheet = self._get_or_create_sheet(workbook, sheet_name, date)

        # Zápis dat
        row = self._find_row_for_date(sheet, date)
        if row:  # Kontrola, zda byl nalezen řádek pro dané datum
            sheet.cell(row=row, column=5, value=start_time)  # Sloupec E
            sheet.cell(row=row, column=6, value=lunch_duration)     # Sloupec F
            sheet.cell(row=row, column=7, value=end_time)    # Sloupec G

            workbook.save(self.file_path)

    def _load_or_create_workbook(self):
        if os.path.exists(self.file_path):
            return load_workbook(self.file_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # Odstranění výchozího listu
            return wb

    def _get_sheet_name(self, date):
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        return f"{date_obj.strftime('%m')}hod{date_obj.strftime('%y')}"

    def _get_or_create_sheet(self, workbook, sheet_name, date):
        if sheet_name not in workbook.sheetnames:
            # Vytvoření nového listu
            if "MMhodRR" in workbook.sheetnames:
                template = workbook["MMhodRR"]
                new_sheet = workbook.copy_worksheet(template)
                new_sheet.title = sheet_name
            else:
                new_sheet = workbook.create_sheet(sheet_name)

            # Úprava vzorců
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            prev_month = (date_obj.replace(day=1) - datetime.timedelta(days=1)).strftime("%m")
            prev_year = (date_obj.replace(day=1) - datetime.timedelta(days=1)).strftime("%y")
            prev_sheet_name = f"{prev_month}hod{prev_year}"

            new_sheet['T3'] = f"='{prev_sheet_name}'!T6"
            new_sheet['Q3'] = f"='{prev_sheet_name}'!Q6"
            new_sheet['O29'] = f"='{prev_sheet_name}'!O27"

            # Doplnění datumů pro celý měsíc
            self._fill_dates(new_sheet, date)

            return new_sheet
        else:
            return workbook[sheet_name]

    def _fill_dates(self, sheet, date):
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        _, last_day = calendar.monthrange(date_obj.year, date_obj.month)
        for day in range(1, last_day + 1):
            current_date = date_obj.replace(day=day)
            sheet.cell(row=day + 2, column=1, value=current_date)

    def _find_row_for_date(self, sheet, date):
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        for row in range(3, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=1).value
            if isinstance(cell_value, datetime) and cell_value.date() == date_obj.date():
                return row
        return None  # Vrácení None, pokud řádek pro dané datum neexistuje

    def update_project_info(self, project_name, start_date, end_date):
        workbook = self._load_or_create_workbook()
        sheet = workbook.active

        # Zápis informací o projektu
        sheet['A1'] = project_name
        sheet['A2'] = start_date
        sheet['A3'] = end_date

        workbook.save(self.file_path)