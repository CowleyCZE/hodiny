"""Správa souboru Hodiny2025.xlsx – měsíční evidence (1 sheet = 1 měsíc).

Zjednodušené schema listu:
 A=Den, B=Datum, C=Den v týdnu, D=Svátek, E/F/G=Začátek/Oběd/Konec,
 H=Celkem hodin (vzorec), I=Přesčasy, M=Počet zaměstnanců, N=Celkem * M.

Třída zajišťuje:
 - lazy vytvoření pracovního sešitu + template list
 - generování / inicializaci listu pro měsíc (01hod25 ... 12hod25)
 - zápis denních údajů + udržení vzorců
 - načítání souhrnů, denních záznamů a validaci integrity
"""
import calendar
import json
import logging
from datetime import datetime, time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet

from config import Config

try:
    from utils.logger import setup_logger

    logger = setup_logger("hodiny2025_manager")
except ImportError:
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger("hodiny2025_manager")


class Hodiny2025Manager:
    """
    Manager pro správu Excel souboru s evidencí pracovních hodin pro rok 2025.
    """

    HEADER_ROW, DATA_START_ROW, DATA_END_ROW, SUMMARY_ROW = 2, 3, 33, 34
    COL_DAY, COL_DATE, COL_WEEKDAY, COL_HOLIDAY = 1, 2, 3, 4
    COL_START, COL_LUNCH, COL_END = 5, 6, 7
    COL_TOTAL_HOURS, COL_OVERTIME, COL_NIGHT, COL_WEEKEND = 8, 9, 10, 11
    COL_HOLIDAY_HOURS, COL_EMPLOYEES, COL_TOTAL_ALL = 12, 13, 14

    CZECH_MONTHS = {
        1: "Leden", 2: "Únor", 3: "Březen", 4: "Duben", 5: "Květen", 6: "Červen",
        7: "Červenec", 8: "Srpen", 9: "Září", 10: "Říjen", 11: "Listopad", 12: "Prosinec",
    }
    CZECH_WEEKDAYS = ["Po", "Út", "St", "Čt", "Pá", "So", "Ne"]

    def __init__(self, excel_path):
        self.excel_path = Path(excel_path)
        self.workbook_name = "Hodiny2025.xlsx"
        self.template_sheet_name = "MMhod25"
        self.file_path = self.excel_path / self.workbook_name
        self._ensure_excel_file_exists()
        logger.info("Hodiny2025Manager inicializován pro soubor: %s", self.file_path)

    def _load_dynamic_config(self):
        """Načte dynamickou konfiguraci pro ukládání do XLSX souborů."""
        from config import Config
        if not Config.CONFIG_FILE_PATH.exists():
            return {}
        try:
            with open(Config.CONFIG_FILE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError) as e:
            logger.error("Chyba při načítání dynamické konfigurace: %s", e, exc_info=True)
            return {}

    def _get_cell_coordinates(self, field_key, sheet_name=None):
        """Vrátí seznam (row, col) souřadnic pro daný field z dynamické konfigurace.
        
        Args:
            field_key: Klíč pole z konfigurace (např. 'start_time', 'date')
            sheet_name: Název listu, pokud chceme ověřit shodu
            
        Returns:
            list: Seznam (row, col) souřadnic nebo prázdný seznam pokud není nakonfigurováno
        """
        config = self._load_dynamic_config()
        monthly_config = config.get('monthly_time', {})
        field_configs = monthly_config.get(field_key, [])
        
        if not field_configs:
            return []
            
        coordinates = []
        for field_config in field_configs:
            # Ověř, že konfigurace je pro správný soubor a list
            if field_config.get('file') != self.workbook_name:
                logger.warning("Konfigurace pro monthly_time/%s odkazuje na jiný soubor: %s", 
                             field_key, field_config.get('file'))
                continue
                
            if sheet_name and field_config.get('sheet') != sheet_name:
                logger.warning("Konfigurace pro monthly_time/%s odkazuje na jiný list: %s (očekáván %s)", 
                             field_key, field_config.get('sheet'), sheet_name)
                continue
                
            cell = field_config.get('cell')
            if not cell:
                continue
                
            try:
                coordinates.append(coordinate_to_tuple(cell))  # Převede např. "A1" na (1, 1)
            except ValueError as e:
                logger.error("Neplatná buňka v konfiguraci pro monthly_time/%s: %s - %s", 
                            field_key, cell, e)
                continue
                
        return coordinates

    def _ensure_excel_file_exists(self):
        if not self.file_path.exists():
            logger.info("Vytváří se nový soubor: %s", self.file_path)
            self._create_new_workbook()

    def _create_new_workbook(self):
        workbook = Workbook()
        if workbook.active:
            workbook.remove(workbook.active)
        template_sheet = workbook.create_sheet(title=self.template_sheet_name)
        self._setup_template_sheet(template_sheet)
        current_month = datetime.now().month
        current_sheet = workbook.copy_worksheet(template_sheet)
        current_sheet.title = f"{current_month:02d}hod25"
        self._setup_month_sheet(current_sheet, current_month, 2025)
        workbook.save(self.file_path)
        logger.info("Vytvořen nový Excel soubor: %s", self.file_path)

    def _setup_template_sheet(self, sheet: Worksheet):
        sheet["A1"] = "Měsíční výkaz práce - [Měsíc] 2025"
        headers = [
            "Den", "Datum", "Den v týdnu", "Svátek", "Začátek", "Oběd (h)", "Konec",
            "Celkem hodin", "Přesčasy", "Noční práce", "Víkend", "Svátky",
            "Zaměstnanci", "Celkem odpracováno",
        ]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=self.HEADER_ROW, column=col)
            if not isinstance(cell, MergedCell):
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        for day in range(1, 32):
            row = self.DATA_START_ROW + day - 1
            sheet.cell(row=row, column=self.COL_DAY).value = day
            formula = f'=IF(AND(E{row}<>"",G{row}<>""),(G{row}-E{row})*24-F{row},0)'
            self._set_cell_formula(sheet, row, self.COL_TOTAL_HOURS, formula)
            self._set_cell_formula(sheet, row, self.COL_OVERTIME, f"=MAX(0,H{row}-8)")
            self._set_cell_formula(sheet, row, self.COL_TOTAL_ALL, f"=H{row}*M{row}")

        self._set_summary_formulas(sheet)

    def _set_summary_formulas(self, sheet: Worksheet):
        self._set_cell_formula(sheet, self.SUMMARY_ROW, 1, "SOUHRN:")
        for col, formula_col in [
            (self.COL_TOTAL_HOURS, "H"), (self.COL_OVERTIME, "I"), (self.COL_TOTAL_ALL, "N")
        ]:
            formula = f"=SUM({formula_col}{self.DATA_START_ROW}:{formula_col}{self.DATA_END_ROW})"
            self._set_cell_formula(sheet, self.SUMMARY_ROW, col, formula)
            sheet.cell(row=self.SUMMARY_ROW, column=col).font = Font(bold=True)
            sheet.cell(row=self.SUMMARY_ROW, column=col).fill = PatternFill("solid", fgColor="CCCCCC")

    def _setup_month_sheet(self, sheet: Worksheet, month: int, year: int):
        month_name = self.CZECH_MONTHS[month]
        sheet.cell(row=1, column=1).value = f"Měsíční výkaz práce - {month_name} {year}"
        days_in_month = calendar.monthrange(year, month)[1]

        for day in range(1, days_in_month + 1):
            row = self.DATA_START_ROW + day - 1
            date_obj = datetime(year, month, day)
            sheet.cell(row=row, column=self.COL_DATE).value = date_obj.strftime("%d.%m.%Y")
            sheet.cell(row=row, column=self.COL_WEEKDAY).value = self.CZECH_WEEKDAYS[date_obj.weekday()]
            if date_obj.weekday() >= 5:
                for col in range(1, 15):
                    sheet.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFE6E6")

        for day in range(days_in_month + 1, 32):
            row = self.DATA_START_ROW + day - 1
            for col in range(1, 15):
                self._set_cell_formula(sheet, row, col, "")

    def get_or_create_month_sheet(self, month: int, year: int = 2025) -> tuple[Workbook, Worksheet]:
        sheet_name = f"{month:02d}hod{str(year)[2:]}"
        try:
            workbook = load_workbook(self.file_path)
        except (FileNotFoundError, InvalidFileException):
            self._create_new_workbook()
            workbook = load_workbook(self.file_path)

        if sheet_name not in workbook.sheetnames:
            if self.template_sheet_name not in workbook.sheetnames:
                raise ValueError(f"Template list '{self.template_sheet_name}' nebyl nalezen")
            template_sheet = workbook[self.template_sheet_name]
            new_sheet = workbook.copy_worksheet(template_sheet)
            new_sheet.title = sheet_name
            self._setup_month_sheet(new_sheet, month, year)
            logger.info("Vytvořen nový list: %s", sheet_name)
            return workbook, new_sheet

        return workbook, workbook[sheet_name]

    def zapis_pracovni_doby(self, date_str, start_time_str, end_time_str, lunch_duration_str, num_employees):
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            workbook, sheet = self.get_or_create_month_sheet(date_obj.month, date_obj.year)
            row = self.DATA_START_ROW + date_obj.day - 1

            self._update_day_record(sheet, row, start_time_str, end_time_str, lunch_duration_str, num_employees)
            self._ensure_formulas_are_set(sheet, row)

            workbook.save(self.file_path)
            logger.info("Pracovní doba pro %s byla zapsána do listu %s", date_str, sheet.title)
        except (ValueError, IOError, FileNotFoundError) as e:
            logger.error("Chyba při zápisu pracovní doby pro %s: %s", date_str, e, exc_info=True)
            raise

    def _update_day_record(self, sheet, row, start_time_str, end_time_str, lunch_duration_str, num_employees):
        # Zápis času začátku s dynamickou konfigurací
        start_time_coords = self._get_cell_coordinates('start_time', sheet.title)
        if start_time_coords and start_time_str and start_time_str != "00:00":
            for start_row, start_col in start_time_coords:
                actual_row = start_row if start_row != row else row
                self._set_cell_value(sheet, actual_row, start_col, datetime.strptime(start_time_str, "%H:%M").time())
                logger.info("Čas začátku zapsán do buňky %s%d (dynamická konfigurace)", 
                           chr(64 + start_col), actual_row)
        elif start_time_str and start_time_str != "00:00":
            # Fallback na původní logiku
            self._set_cell_value(sheet, row, self.COL_START, datetime.strptime(start_time_str, "%H:%M").time())
            
        # Zápis času konce s dynamickou konfigurací
        end_time_coords = self._get_cell_coordinates('end_time', sheet.title)
        if end_time_coords and end_time_str and end_time_str != "00:00":
            for end_row, end_col in end_time_coords:
                actual_row = end_row if end_row != row else row
                self._set_cell_value(sheet, actual_row, end_col, datetime.strptime(end_time_str, "%H:%M").time())
                logger.info("Čas konce zapsán do buňky %s%d (dynamická konfigurace)", 
                           chr(64 + end_col), actual_row)
        elif end_time_str and end_time_str != "00:00":
            # Fallback na původní logiku
            self._set_cell_value(sheet, row, self.COL_END, datetime.strptime(end_time_str, "%H:%M").time())

        # Zápis doby oběda s dynamickou konfigurací
        lunch_hours = float(lunch_duration_str) if lunch_duration_str else 0.0
        lunch_coords = self._get_cell_coordinates('lunch_hours', sheet.title)
        if lunch_coords:
            for lunch_row, lunch_col in lunch_coords:
                actual_row = lunch_row if lunch_row != row else row
                lunch_cell = self._set_cell_value(sheet, actual_row, lunch_col, lunch_hours)
                if lunch_cell:
                    lunch_cell.number_format = "0.0"
                logger.info("Doba oběda zapsána do buňky %s%d (dynamická konfigurace)", 
                           chr(64 + lunch_col), actual_row)
        else:
            # Fallback na původní logiku
            lunch_cell = self._set_cell_value(sheet, row, self.COL_LUNCH, lunch_hours)
            if lunch_cell:
                lunch_cell.number_format = "0.0"

        # Zápis počtu zaměstnanců s dynamickou konfigurací
        employees_coords = self._get_cell_coordinates('num_employees', sheet.title)
        if employees_coords:
            for emp_row, emp_col in employees_coords:
                actual_row = emp_row if emp_row != row else row
                self._set_cell_value(sheet, actual_row, emp_col, num_employees if num_employees > 0 else 0)
                logger.info("Počet zaměstnanců zapsán do buňky %s%d (dynamická konfigurace)", 
                           chr(64 + emp_col), actual_row)
        else:
            # Fallback na původní logiku
            self._set_cell_value(sheet, row, self.COL_EMPLOYEES, num_employees if num_employees > 0 else 0)

    def _ensure_formulas_are_set(self, sheet, row):
        formulas = {
            self.COL_TOTAL_HOURS: f'=IF(AND(E{row}<>"",G{row}<>""),(G{row}-E{row})*24-F{row},0)',
            self.COL_OVERTIME: f"=MAX(0,H{row}-8)",
            self.COL_TOTAL_ALL: f"=H{row}*M{row}",
        }
        for col, formula in formulas.items():
            cell = sheet.cell(row=row, column=col)
            if not isinstance(cell, MergedCell) and (not cell.value or not str(cell.value).startswith("=")):
                cell.value = formula

    def get_monthly_summary(self, month: int, year: int = 2025) -> dict:
        summary = {
            "month": month, "year": year, "month_name": self.CZECH_MONTHS.get(month, "Neznámý"),
            "total_hours": 0, "total_overtime": 0, "total_all_employees": 0,
            "sheet_name": f"{month:02d}hod{str(year)[2:]}", "error": None,
        }
        try:
            _, sheet = self.get_or_create_month_sheet(month, year)
            summary.update({
                "total_hours": self._safe_float(sheet.cell(self.SUMMARY_ROW, self.COL_TOTAL_HOURS).value),
                "total_overtime": self._safe_float(sheet.cell(self.SUMMARY_ROW, self.COL_OVERTIME).value),
                "total_all_employees": self._safe_float(sheet.cell(self.SUMMARY_ROW, self.COL_TOTAL_ALL).value),
            })
        except (ValueError, IOError, FileNotFoundError) as e:
            logger.error("Chyba při získávání měsíčního souhrnu pro %d/%d: %s", month, year, e)
            summary["error"] = str(e)
        return summary

    def get_daily_record(self, date_str: str) -> dict:
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            _, sheet = self.get_or_create_month_sheet(date_obj.month, date_obj.year)
            row = self.DATA_START_ROW + date_obj.day - 1

            # Load with data_only=True to get calculated values
            data_workbook = load_workbook(self.file_path, data_only=True)
            data_sheet = data_workbook[sheet.title]

            record = self._extract_daily_data(data_sheet, row)
            record["date"] = date_str
            record["day"] = date_obj.day
            record["row"] = row
            record["sheet_name"] = sheet.title

            self._recalculate_if_needed(record)

            return record
        except (ValueError, IOError, FileNotFoundError) as e:
            logger.error("Chyba při získávání záznamu pro %s: %s", date_str, e)
            return {"date": date_str, "error": str(e)}

    def _extract_daily_data(self, sheet, row):
        return {
            "start_time": self._safe_time_format(sheet.cell(row, self.COL_START).value),
            "end_time": self._safe_time_format(sheet.cell(row, self.COL_END).value),
            "lunch_hours": self._safe_float(sheet.cell(row, self.COL_LUNCH).value),
            "total_hours": self._safe_float(sheet.cell(row, self.COL_TOTAL_HOURS).value),
            "overtime": self._safe_float(sheet.cell(row, self.COL_OVERTIME).value),
            "num_employees": self._safe_int(sheet.cell(row, self.COL_EMPLOYEES).value),
            "total_all_employees": self._safe_float(sheet.cell(row, self.COL_TOTAL_ALL).value),
        }

    def _recalculate_if_needed(self, record):
        if record["total_hours"] == 0.0 and record["start_time"] and record["end_time"]:
            try:
                start = datetime.strptime(record["start_time"], "%H:%M")
                end = datetime.strptime(record["end_time"], "%H:%M")
                hours = (end - start).total_seconds() / 3600 - record["lunch_hours"]
                record["total_hours"] = max(0.0, hours)
            except (ValueError, TypeError):
                pass

        if record["overtime"] == 0.0 and record["total_hours"] > 8.0:
            record["overtime"] = record["total_hours"] - 8.0

        if record["total_all_employees"] == 0.0 and record["total_hours"] > 0:
            record["total_all_employees"] = record["total_hours"] * record["num_employees"]

    def _set_cell_value(self, sheet, row, col, value):
        cell = sheet.cell(row=row, column=col)
        if not isinstance(cell, MergedCell):
            cell.value = value
            return cell
        return None

    def _set_cell_formula(self, sheet, row, col, formula):
        return self._set_cell_value(sheet, row, col, formula)

    def _safe_time_format(self, value):
        if isinstance(value, time):
            return value.strftime("%H:%M")
        return value if isinstance(value, str) else None

    def _safe_float(self, value):
        try:
            return float(value) if value is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    def _safe_int(self, value):
        try:
            return int(value) if value is not None else 0
        except (ValueError, TypeError):
            return 0

    def create_test_data(self):
        logger.info("Vytváří se testovací data pro Hodiny2025.xlsx")
        test_dates = [
            ("2025-01-02", "07:00", "15:30", "0.5", 3),
            ("2025-01-03", "07:00", "16:00", "1.0", 3),
            ("2025-01-06", "08:00", "16:30", "0.5", 2),
        ]
        for data in test_dates:
            try:
                self.zapis_pracovni_doby(*data)
                logger.info("✅ Testovací záznam vytvořen: %s", data[0])
            except Exception as e:
                logger.error("❌ Chyba při vytváření testovacího záznamu %s: %s", data[0], e)

    def validate_data_integrity(self) -> dict:
        results = {"valid": True, "errors": [], "warnings": [], "sheets_checked": 0, "records_checked": 0}
        try:
            workbook = load_workbook(self.file_path)
            sheet_names = [s for s in workbook.sheetnames if s != self.template_sheet_name]
            results["sheets_checked"] = len(sheet_names)

            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                for row in range(self.DATA_START_ROW, self.DATA_END_ROW + 1):
                    results["records_checked"] += 1
                    total_formula = sheet.cell(row=row, column=self.COL_TOTAL_HOURS).value
                    if isinstance(total_formula, str) and not total_formula.startswith("="):
                        results["valid"] = False
                        results["errors"].append(f"List {sheet_name}, řádek {row}: Chybí vzorec.")

                    start = sheet.cell(row=row, column=self.COL_START).value
                    end = sheet.cell(row=row, column=self.COL_END).value
                    if bool(start) != bool(end):
                        results["warnings"].append(f"List {sheet_name}, řádek {row}: Chybí čas začátku/konce.")
        except Exception as e:
            results["valid"] = False
            results["errors"].append(f"Chyba při validaci: {e}")
        return results
