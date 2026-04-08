"""Služby pro práci s Excel soubory používané advanced konfigurací."""

from openpyxl import load_workbook

from config import Config
from utils.logger import setup_logger

logger = setup_logger("excel_file_service")


def list_excel_files():
    """Vrátí seřazený seznam všech dostupných XLSX souborů."""
    return sorted(path.name for path in Config.EXCEL_BASE_PATH.glob("*.xlsx"))


def get_sheet_names(filename):
    """Vrátí názvy listů v zadaném Excel souboru."""
    file_path = Config.EXCEL_BASE_PATH / filename
    if not file_path.exists():
        raise FileNotFoundError("Soubor nenalezen")

    workbook = load_workbook(file_path, read_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def get_sheet_content(filename, sheet_name, max_rows=None, max_cols=26):
    """Vrátí obsah listu ve formátu vhodném pro frontend."""
    file_path = Config.EXCEL_BASE_PATH / filename
    if not file_path.exists():
        raise FileNotFoundError("Soubor nenalezen")

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError("List nenalezen")

        sheet = workbook[sheet_name]
        row_limit = min(sheet.max_row, max_rows or Config.MAX_ROWS_TO_DISPLAY_EXCEL_VIEWER)
        col_limit = min(sheet.max_column, max_cols)

        data = []
        for row_idx in range(1, row_limit + 1):
            row_data = []
            for col_idx in range(1, col_limit + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            data.append(row_data)

        return {"data": data, "rows": row_limit, "cols": col_limit}
    finally:
        workbook.close()


def rename_excel_file(old_filename, new_filename):
    """Přejmenuje XLSX soubor po základní validaci vstupu."""
    if not old_filename or not new_filename:
        raise ValueError("Chybí název souboru")

    if not old_filename.endswith(".xlsx") or not new_filename.endswith(".xlsx"):
        raise ValueError("Pouze .xlsx soubory mohou být přejmenovány")

    old_path = Config.EXCEL_BASE_PATH / old_filename
    new_path = Config.EXCEL_BASE_PATH / new_filename

    if not old_path.exists():
        raise FileNotFoundError(f"Soubor {old_filename} neexistuje")

    if new_path.exists():
        raise FileExistsError(f"Soubor {new_filename} již existuje")

    old_path.rename(new_path)
    logger.info("Soubor %s byl přejmenován na %s", old_filename, new_filename)
    return new_filename
