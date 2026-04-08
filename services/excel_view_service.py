"""Služby pro čtení Excel souborů pro viewer a editor."""

import datetime as dt

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from config import Config
from services.excel_file_service import list_excel_files


def get_excel_viewer_context(excel_manager, requested_file=None, active_sheet_name=None, selected_category=None):
    """Připraví kontext pro read-only Excel viewer."""
    base_path = Config.EXCEL_BASE_PATH
    all_files = list_excel_files()
    all_metadata = excel_manager.get_all_metadata()
    categories = sorted(list(set(meta.get("category", "Ostatní") for meta in all_metadata.values())))

    if selected_category and selected_category != "Vše":
        excel_files = [name for name in all_files if all_metadata.get(name, {}).get("category") == selected_category]
    else:
        excel_files = all_files

    if not excel_files:
        return {
            "excel_files": [],
            "selected_file": None,
            "sheet_names": [],
            "active_sheet": None,
            "data": [],
            "all_metadata": all_metadata,
            "categories": categories,
            "selected_category": selected_category,
        }

    selected_file = requested_file if requested_file in excel_files else excel_files[0]
    selected_path = base_path / selected_file

    try:
        workbook = load_workbook(selected_path, read_only=True, data_only=True)
    except (FileNotFoundError, InvalidFileException) as exc:
        raise FileNotFoundError(f"Chyba při zobrazení souboru: {exc}") from exc

    try:
        sheet_names = workbook.sheetnames
        if not sheet_names:
            return {
                "excel_files": excel_files,
                "selected_file": selected_file,
                "sheet_names": [],
                "active_sheet": None,
                "data": [],
                "all_metadata": all_metadata,
                "categories": categories,
                "selected_category": selected_category,
            }

        active_sheet = active_sheet_name if active_sheet_name in sheet_names else sheet_names[0]
        sheet = workbook[active_sheet]
        data = []

        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= Config.MAX_ROWS_TO_DISPLAY_EXCEL_VIEWER:
                break
            data.append(_format_viewer_row(row))
    finally:
        workbook.close()

    return {
        "excel_files": excel_files,
        "selected_file": selected_file,
        "sheet_names": sheet_names,
        "active_sheet": active_sheet,
        "data": data,
        "all_metadata": all_metadata,
        "categories": categories,
        "selected_category": selected_category,
    }


def get_excel_editor_context(active_filename, requested_file=None, active_sheet_name=None):
    """Připraví kontext pro editovatelný Excel viewer."""
    base_path = Config.EXCEL_BASE_PATH
    excel_files = list_excel_files()
    if not excel_files:
        return {
            "excel_files": [],
            "selected_file": None,
            "sheet_names": [],
            "active_sheet": None,
            "data": [],
        }

    selected_file = requested_file if requested_file in excel_files else active_filename
    selected_path = base_path / selected_file

    try:
        workbook = load_workbook(selected_path, data_only=False)
    except (FileNotFoundError, InvalidFileException) as exc:
        raise FileNotFoundError(f"Chyba při zobrazení souboru: {exc}") from exc

    try:
        sheet_names = workbook.sheetnames
        if not sheet_names:
            return {
                "excel_files": excel_files,
                "selected_file": selected_file,
                "sheet_names": [],
                "active_sheet": None,
                "data": [],
            }

        active_sheet = active_sheet_name if active_sheet_name in sheet_names else sheet_names[0]
        sheet = workbook[active_sheet]

        data_with_coords = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_idx > Config.MAX_ROWS_TO_DISPLAY_EXCEL_VIEWER:
                break
            row_data = []
            for col_idx, cell_value in enumerate(row, 1):
                row_data.append(
                    {"value": str(cell_value) if cell_value is not None else "", "row": row_idx, "col": col_idx}
                )
            data_with_coords.append(row_data)
    finally:
        workbook.close()

    return {
        "excel_files": excel_files,
        "selected_file": selected_file,
        "sheet_names": sheet_names,
        "active_sheet": active_sheet,
        "data": data_with_coords,
    }


def _format_viewer_row(row):
    formatted_row = []
    for cell in row:
        if cell is None:
            formatted_row.append("")
        elif isinstance(cell, (dt.datetime, dt.date)):
            if isinstance(cell, dt.datetime) and cell.hour == 0 and cell.minute == 0 and cell.second == 0:
                formatted_row.append(cell.strftime("%d.%m.%Y"))
            else:
                formatted_row.append(cell.strftime("%d.%m.%Y %H:%M"))
        elif isinstance(cell, dt.time):
            formatted_row.append(cell.strftime("%H:%M"))
        elif isinstance(cell, float):
            formatted_row.append(str(int(cell)) if cell.is_integer() else f"{cell:.2f}")
        else:
            formatted_row.append(str(cell))
    return formatted_row
