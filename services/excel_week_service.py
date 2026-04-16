"""Helpery pro týdenní Excel soubory a zápis docházky."""

import shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from config import Config
from utils.logger import setup_logger

logger = setup_logger("excel_week_service")


def _resolve_base_coordinate(get_cell_coordinates, field_key, sheet_name, data_type):
    coordinates = get_cell_coordinates(field_key, sheet_name, data_type)
    return coordinates[0] if coordinates else None


def _resolve_weekday_column(base_coordinate, weekday):
    if not base_coordinate:
        return None
    _, base_column = base_coordinate
    return base_column + (weekday * 2)


def _resolve_employee_anchor(get_cell_coordinates, sheet_name):
    configured_employee_cell = _resolve_base_coordinate(
        get_cell_coordinates,
        "employee_name",
        sheet_name,
        "weekly_time",
    )
    if configured_employee_cell:
        return configured_employee_cell
    return Config.EXCEL_EMPLOYEE_START_ROW, 1


def _get_or_create_employee_row(sheet, employee_name, employee_start_row, employee_name_column):
    for row_index in range(employee_start_row, sheet.max_row + 1):
        existing_name = sheet.cell(row=row_index, column=employee_name_column).value
        if existing_name == employee_name:
            return row_index
        if existing_name in (None, ""):
            sheet.cell(row=row_index, column=employee_name_column, value=employee_name)
            return row_index

    row_index = max(sheet.max_row + 1, employee_start_row)
    sheet.cell(row=row_index, column=employee_name_column, value=employee_name)
    return row_index


def _write_sheet_cell(sheet, row_index, column_index, value, number_format=None):
    target_cell = sheet.cell(row=row_index, column=column_index)
    if isinstance(target_cell, MergedCell):
        return

    target_cell.value = value
    if number_format:
        target_cell.number_format = number_format


def archive_active_week_file(active_file_path, workbook, current_week_number, last_archived_week):
    """Archivuje aktivní soubor a vyčistí týdenní listy v aktivním workbooku."""
    archive_filename = f"Hodiny_Cap_Tyden_{last_archived_week}.xlsx"
    archive_path = active_file_path.parent / archive_filename
    shutil.copy(active_file_path, archive_path)
    logger.info("Archivován soubor: %s", archive_path)

    for sheet_name in list(workbook.sheetnames):
        if sheet_name.startswith(Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME):
            workbook.remove(workbook[sheet_name])

    workbook.create_sheet(Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME)


def get_or_create_weekly_file(base_path, active_file_path, week_number):
    """Vrátí cestu k týdennímu souboru a případně ho vytvoří."""
    weekly_filename = f"{active_file_path.stem}_Tyden{week_number}.xlsx"
    weekly_file_path = base_path / weekly_filename

    if weekly_file_path.exists():
        return weekly_file_path

    previous_week_file = find_previous_weekly_file(base_path, active_file_path.stem, week_number)
    if previous_week_file and previous_week_file.exists():
        shutil.copy(previous_week_file, weekly_file_path)
        logger.info("Vytvořen týdenní soubor %s zkopírováním z %s", weekly_filename, previous_week_file.name)
    else:
        shutil.copy(active_file_path, weekly_file_path)
        logger.info("Vytvořen týdenní soubor %s zkopírováním ze šablony %s", weekly_filename, active_file_path.name)

    return weekly_file_path


def find_previous_weekly_file(base_path, active_stem, current_week):
    """Najde nejbližší předchozí týdenní soubor."""
    for week in range(current_week - 1, 0, -1):
        potential_file = base_path / f"{active_stem}_Tyden{week}.xlsx"
        if potential_file.exists():
            return potential_file
    return None


def open_weekly_workbook(weekly_file_path):
    """Otevře týdenní workbook pro zápis."""
    try:
        return load_workbook(str(weekly_file_path))
    except Exception as exc:
        logger.error("Chyba při otevírání týdenního souboru %s: %s", weekly_file_path, exc, exc_info=True)
        return None


def ensure_week_sheet(workbook, week_number):
    """Vrátí existující týdenní list nebo ho vytvoří ze šablony."""
    week_sheet_name = (
        Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME
        if Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME in workbook.sheetnames
        else f"{Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME} {week_number}"
    )

    if week_sheet_name not in workbook.sheetnames:
        create_week_sheet_from_template(workbook, week_sheet_name)

    return week_sheet_name, workbook[week_sheet_name]


def create_week_sheet_from_template(workbook, sheet_name):
    """Vytvoří nový list zkopírováním šablony `Týden` v rámci workbooku."""
    try:
        template_name = Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME
        if template_name in workbook.sheetnames:
            source = workbook[template_name]
            target = workbook.copy_worksheet(source)
            target.title = sheet_name
            logger.info("Vytvořen list %s zkopírováním šablony %s", sheet_name, template_name)
        else:
            workbook.create_sheet(sheet_name)
            logger.warning("Šablona %s nenalezena, vytvořen prázdný list %s", template_name, sheet_name)
    except Exception as exc:
        logger.error("Chyba při kopírování listu: %s", exc, exc_info=True)
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)


def write_time_entry_to_sheet(
    sheet,
    sheet_name,
    date_obj,
    start_time_str,
    end_time_str,
    lunch_duration_str,
    employees,
    get_cell_coordinates,
    current_project_name,
    work_description="",
    category="",
):
    """Zapíše docházku do konkrétního listu."""
    logger.info("Zápis docházky: AI Kategorie: %s, Popis práce: %s", category, work_description)
    weekday = date_obj.weekday()
    total_hours_base = _resolve_base_coordinate(get_cell_coordinates, "total_hours", sheet_name, "weekly_time")
    start_time_base = _resolve_base_coordinate(get_cell_coordinates, "start_time", sheet_name, "weekly_time")
    end_time_base = _resolve_base_coordinate(get_cell_coordinates, "end_time", sheet_name, "weekly_time")
    date_base = _resolve_base_coordinate(get_cell_coordinates, "date", sheet_name, "weekly_time")
    employee_start_row, employee_name_column = _resolve_employee_anchor(get_cell_coordinates, sheet_name)

    hours_column = _resolve_weekday_column(total_hours_base, weekday)
    start_column = _resolve_weekday_column(start_time_base, weekday)
    end_column = _resolve_weekday_column(end_time_base, weekday)
    date_column = _resolve_weekday_column(date_base, weekday)

    if hours_column is None:
        hours_column = 2 + (weekday * 2)
    if start_column is None:
        start_column = 2 + (weekday * 2)
    if end_column is None:
        end_column = 3 + (weekday * 2)
    if date_column is None:
        date_column = 2 + (weekday * 2)

    start_time = datetime.strptime(start_time_str, "%H:%M")
    end_time = datetime.strptime(end_time_str, "%H:%M")
    total_hours = round((end_time - start_time).total_seconds() / 3600 - float(lunch_duration_str), 2)
    write_times = start_time_str != "00:00" or end_time_str != "00:00"

    for employee in employees:
        row_index = _get_or_create_employee_row(sheet, employee, employee_start_row, employee_name_column)
        _write_sheet_cell(sheet, row_index, hours_column, total_hours, number_format="0.00")

    if start_time_base:
        _write_sheet_cell(sheet, start_time_base[0], start_column, start_time_str if write_times else None)
    else:
        _write_sheet_cell(sheet, 7, start_column, start_time_str if write_times else None)

    if end_time_base:
        _write_sheet_cell(sheet, end_time_base[0], end_column, end_time_str if write_times else None)
    else:
        _write_sheet_cell(sheet, 7, end_column, end_time_str if write_times else None)

    if date_base:
        _write_sheet_cell(sheet, date_base[0], date_column, date_obj.date(), number_format="DD.MM.YYYY")
    else:
        _write_sheet_cell(sheet, 80, date_column, date_obj.date(), number_format="DD.MM.YYYY")

    project_coords = get_cell_coordinates("project_name", sheet_name, "projects")
    if current_project_name and project_coords:
        for project_row, project_col in project_coords:
            _write_sheet_cell(sheet, project_row, project_col, f"NÁZEV PROJEKTU : {current_project_name}")


def get_current_week_preview(workbook, week_number):
    """Vrátí omezený náhled dat pro aktuální nebo zadaný týden."""
    sheet_name_dynamic = f"{Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME} {week_number}"
    if sheet_name_dynamic in workbook.sheetnames:
        sheet = workbook[sheet_name_dynamic]
        sheet_display_name = sheet_name_dynamic
    elif Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME in workbook.sheetnames:
        sheet = workbook[Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME]
        sheet_display_name = Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME
    else:
        return None

    data = []
    max_row = min(sheet.max_row, 20)
    max_col = min(sheet.max_column, 10)

    for row_index in range(1, max_row + 1):
        row_data = []
        for col_index in range(1, max_col + 1):
            value = sheet.cell(row=row_index, column=col_index).value
            if value is None:
                value = ""
            elif isinstance(value, (int, float)):
                value = str(value)
            row_data.append(str(value))
        data.append(row_data)

    return {
        "sheet_name": sheet_display_name,
        "data": data,
        "rows": len(data),
        "cols": len(data[0]) if data else 0,
    }
