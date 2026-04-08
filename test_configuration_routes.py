import pytest
from openpyxl import Workbook

from app import app
from config import Config


@pytest.fixture
def isolated_client(tmp_path, monkeypatch):
    data_path = tmp_path / "data"
    excel_path = tmp_path / "excel"
    settings_path = data_path / "settings.json"
    config_path = tmp_path / "config.json"

    data_path.mkdir(parents=True, exist_ok=True)
    excel_path.mkdir(parents=True, exist_ok=True)

    monkeypatch.setattr(Config, "DATA_PATH", data_path)
    monkeypatch.setattr(Config, "EXCEL_BASE_PATH", excel_path)
    monkeypatch.setattr(Config, "SETTINGS_FILE_PATH", settings_path)
    monkeypatch.setattr(Config, "CONFIG_FILE_PATH", config_path)

    template_path = excel_path / Config.EXCEL_TEMPLATE_NAME
    workbook = Workbook()
    weekly_sheet = workbook.active
    weekly_sheet.title = Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME
    advances_sheet = workbook.create_sheet(Config.EXCEL_ADVANCES_SHEET_NAME)
    advances_sheet["B80"] = Config.DEFAULT_ADVANCE_OPTION_1
    advances_sheet["D80"] = Config.DEFAULT_ADVANCE_OPTION_2
    advances_sheet["F80"] = Config.DEFAULT_ADVANCE_OPTION_3
    advances_sheet["H80"] = Config.DEFAULT_ADVANCE_OPTION_4
    workbook.save(template_path)

    app.config["TESTING"] = True
    with app.test_client() as client:
        yield client, settings_path, config_path, excel_path


def test_settings_page_is_available(isolated_client):
    client, _, _, _ = isolated_client

    response = client.get("/settings")

    assert response.status_code == 200
    assert "Nastavení aplikace" in response.get_data(as_text=True)


def test_settings_page_persists_runtime_settings(isolated_client):
    client, settings_path, _, _ = isolated_client

    response = client.post(
        "/settings",
        data={
            "project_name": "Projekt Alfa",
            "start_date": "2026-04-01",
            "end_date": "2026-04-30",
            "start_time": "08:15",
            "end_time": "17:30",
            "lunch_duration": "0.5",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    saved_settings = settings_path.read_text(encoding="utf-8")
    assert "Projekt Alfa" in saved_settings
    assert "2026-04-01" in saved_settings
    assert "08:15" in saved_settings
    assert "17:30" in saved_settings


def test_advanced_settings_page_is_available(isolated_client):
    client, _, _, _ = isolated_client

    response = client.get("/nastaveni")

    assert response.status_code == 200
    assert "Rozšířené nastavení ukládání do XLSX souborů" in response.get_data(as_text=True)


def test_dynamic_config_api_roundtrip(isolated_client):
    client, _, _, _ = isolated_client
    payload = {"weekly_time": {"date": [{"file": Config.EXCEL_TEMPLATE_NAME, "sheet": "Týden", "cell": "B6"}]}}

    save_response = client.post("/api/settings", json=payload)
    load_response = client.get("/api/settings")

    assert save_response.status_code == 200
    assert load_response.status_code == 200
    assert load_response.get_json() == payload


def test_excel_file_api_lists_sheets_and_content(isolated_client):
    client, _, _, excel_path = isolated_client
    report_file = excel_path / "report.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Přehled"
    sheet["A1"] = "Test"
    workbook.save(report_file)

    files_response = client.get("/api/files")
    sheets_response = client.get("/api/sheets/report.xlsx")
    content_response = client.get("/api/sheet_content/report.xlsx/P%C5%99ehled")

    assert files_response.status_code == 200
    assert "report.xlsx" in files_response.get_json()["files"]
    assert sheets_response.status_code == 200
    assert sheets_response.get_json()["sheets"] == ["Přehled"]
    assert content_response.status_code == 200
    assert content_response.get_json()["data"][0][0] == "Test"


def test_excel_file_rename_api_renames_file(isolated_client):
    client, _, _, excel_path = isolated_client
    original_path = excel_path / "stary.xlsx"
    Workbook().save(original_path)

    response = client.post(
        "/api/files/rename",
        json={"old_filename": "stary.xlsx", "new_filename": "novy.xlsx"},
    )

    assert response.status_code == 200
    assert not original_path.exists()
    assert (excel_path / "novy.xlsx").exists()
