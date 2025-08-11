import unittest
import tempfile
from pathlib import Path
import datetime
from openpyxl import Workbook
from excel_manager import ExcelManager
from config import Config

class TestExcelManagerReports(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.mock_excel_path_dir = Path(self.temp_dir.name)
        self.mock_excel_filename = "test_report.xlsx"
        self.mock_template_filename = "template.xlsx"
        
        mock_template_path = self.mock_excel_path_dir / self.mock_template_filename
        wb_template = Workbook()
        wb_template.create_sheet(Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME)
        wb_template.save(mock_template_path)

        self.excel_manager = ExcelManager(base_path=self.mock_excel_path_dir,
                                          active_filename=self.mock_excel_filename,
                                          template_filename=self.mock_template_filename)
        
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "Týden 1"
        sheet1['A9'], sheet1['A10'], sheet1['A11'] = "Pepa Novák", "Jana Modrá", "Karel Bílý"
        sheet1['B80'], sheet1['D80'], sheet1['F80'], sheet1['H80'], sheet1['J80'] = (
            datetime.date(2023, 1, 2), datetime.date(2023, 1, 3), datetime.date(2023, 1, 4), 
            datetime.date(2023, 1, 5), datetime.date(2023, 1, 6)
        )
        sheet1['C9'], sheet1['E9'], sheet1['G9'], sheet1['I9'] = 8, 7.5, 0, 6
        sheet1['C10'], sheet1['E10'], sheet1['I10'], sheet1['K10'] = 7, 8, 0, 7.5
        sheet1['C11'] = 5

        sheet2 = wb.create_sheet("Týden 2")
        sheet2['A9'], sheet2['A10'] = "Pepa Novák", "Jana Modrá"
        sheet2['B80'], sheet2['D80'] = datetime.date(2023, 1, 9), datetime.date(2023, 1, 10)
        sheet2['C9'], sheet2['E9'], sheet2['C10'] = 8.5, 8, 6.5
        
        sheet3 = wb.create_sheet("Týden 5")
        sheet3['A9'], sheet3['A10'] = "Pepa Novák", "Karel Bílý"
        sheet3['B80'], sheet3['D80'] = datetime.date(2023, 2, 6), datetime.date(2023, 2, 7)
        sheet3['C9'], sheet3['C10'], sheet3['E10'] = 7, 8, 4

        sheet4 = wb.create_sheet("Týden 6")
        sheet4['A9'] = "Pepa Novák"
        sheet4['B80'], sheet4['D80'] = "neplatne-datum", datetime.date(2023, 1, 17)
        sheet4['C9'], sheet4['E9'] = 5, 5

        sheet_span = wb.create_sheet("Týden Span")
        sheet_span['A9'] = "Test Span User"
        sheet_span['B80'], sheet_span['D80'], sheet_span['F80'] = (
            datetime.date(2023, 1, 30), datetime.date(2023, 1, 31), datetime.date(2023, 2, 1)
        )
        sheet_span['C9'], sheet_span['E9'], sheet_span['G9'] = 8, 7, 6

        wb.save(self.mock_excel_path_dir / self.mock_excel_filename)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_invalid_input_parameters(self):
        with self.assertRaises(ValueError):
            self.excel_manager.generate_monthly_report(month=0, year=2023)
        with self.assertRaises(ValueError):
            self.excel_manager.generate_monthly_report(month=1, year=1999)

    def test_january_report(self):
        report = self.excel_manager.generate_monthly_report(month=1, year=2023)
        self.assertEqual(report.get("Pepa Novák", {}).get("total_hours"), 38.0)
        self.assertEqual(report.get("Pepa Novák", {}).get("free_days"), 1)
        self.assertEqual(report.get("Jana Modrá", {}).get("total_hours"), 29.0)
        self.assertEqual(report.get("Jana Modrá", {}).get("free_days"), 1)
        self.assertEqual(report.get("Karel Bílý", {}).get("total_hours"), 5.0)
        self.assertEqual(report.get("Test Span User", {}).get("total_hours"), 15.0)

    def test_february_report(self):
        report = self.excel_manager.generate_monthly_report(month=2, year=2023)
        self.assertEqual(report.get("Pepa Novák", {}).get("total_hours"), 7.0)
        self.assertEqual(report.get("Karel Bílý", {}).get("total_hours"), 12.0)
        self.assertEqual(report.get("Test Span User", {}).get("total_hours"), 6.0)

    def test_employee_filtering(self):
        report = self.excel_manager.generate_monthly_report(month=1, year=2023, employees=["Pepa Novák", "Jana Modrá"])
        self.assertIn("Pepa Novák", report)
        self.assertIn("Jana Modrá", report)
        self.assertNotIn("Karel Bílý", report)

    def test_no_data_for_month(self):
        report = self.excel_manager.generate_monthly_report(month=3, year=2023)
        self.assertEqual(len(report), 0)

if __name__ == '__main__':
    unittest.main()
