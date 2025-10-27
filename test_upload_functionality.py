"""Tests for file upload and download functionality."""

import os
import tempfile
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app import app
from config import Config


@pytest.fixture
def client():
    """Create test client."""
    app.config["TESTING"] = True
    app.config["WTF_CSRF_ENABLED"] = False
    with app.test_client() as client:
        yield client


@pytest.fixture
def temp_excel_file():
    """Create temporary Excel file for testing."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(title="Sheet1")
        ws["A1"] = "Test data"
        ws["B1"] = "for upload"
        wb.save(tmp.name)
        yield tmp.name
    os.unlink(tmp.name)


def test_upload_new_file(client, temp_excel_file):
    """Test uploading a new Excel file."""
    filename = "test_upload_new.xlsx"

    with open(temp_excel_file, "rb") as f:
        response = client.post("/upload", data={"file": (f, filename)})

    # Should redirect to index
    assert response.status_code == 302

    # Check file was created
    uploaded_path = Config.EXCEL_BASE_PATH / filename
    assert uploaded_path.exists()

    # Clean up
    if uploaded_path.exists():
        uploaded_path.unlink()


def test_upload_invalid_file_extension(client):
    """Test uploading file with invalid extension."""
    with tempfile.NamedTemporaryFile(suffix=".txt", delete=False) as tmp:
        tmp.write(b"Not an Excel file")
        tmp.flush()

        with open(tmp.name, "rb") as f:
            response = client.post("/upload", data={"file": (f, "test.txt")})

    # Should redirect to index
    assert response.status_code == 302

    # Clean up
    os.unlink(tmp.name)


def test_upload_existing_file_shows_confirmation(client, temp_excel_file):
    """Test that uploading existing file shows confirmation page."""
    filename = "test_existing.xlsx"
    upload_path = Config.EXCEL_BASE_PATH / filename

    # Create existing file
    with open(temp_excel_file, "rb") as src:
        with open(upload_path, "wb") as dst:
            dst.write(src.read())

    try:
        with open(temp_excel_file, "rb") as f:
            response = client.post("/upload", data={"file": (f, filename)})

        # Should show confirmation page
        assert response.status_code == 200
        assert "Potvrzení přepsání souboru" in response.get_data(as_text=True)
        assert filename in response.get_data(as_text=True)

    finally:
        # Clean up
        if upload_path.exists():
            upload_path.unlink()


def test_upload_confirm_overwrite(client):
    """Test confirming overwrite of existing file."""
    filename = "test_confirm.xlsx"
    temp_filename = f"temp_{filename}"

    # Create temp file to simulate stored upload
    temp_path = Config.EXCEL_BASE_PATH / temp_filename
    wb = Workbook()
    wb.save(temp_path)

    try:
        response = client.post("/upload/confirm", data={"filename": filename, "temp_filename": temp_filename})

        # Should redirect to index
        assert response.status_code == 302

        # Temp file should be moved to final location
        final_path = Config.EXCEL_BASE_PATH / filename
        assert final_path.exists()
        assert not temp_path.exists()

    finally:
        # Clean up
        final_path = Config.EXCEL_BASE_PATH / filename
        if final_path.exists():
            final_path.unlink()
        if temp_path.exists():
            temp_path.unlink()


def test_download_functionality(client):
    """Test the download route."""
    response = client.get("/download")

    # Should return file or redirect if file not found
    assert response.status_code in [200, 302]


def test_upload_in_memory_and_verify_contents(client):
    """Simulate uploading an in-memory xlsx and verify it was saved with correct contents."""
    filename = "in_memory_test.xlsx"

    # Create workbook in memory
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(title="Sheet1")
    ws["A1"] = "in-memory-test"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Post to upload endpoint
    response = client.post(
        "/upload",
        data={"file": (buf, filename)},
        content_type="multipart/form-data",
        follow_redirects=False,
    )

    # Expect redirect to index after successful upload
    assert response.status_code == 302

    # Verify saved file exists and content matches
    saved_path = Config.EXCEL_BASE_PATH / filename
    assert saved_path.exists()

    try:
        wb_loaded = load_workbook(saved_path, data_only=True)
        ws_loaded = wb_loaded.active if wb_loaded.active is not None else wb_loaded.worksheets[0]
        val = ws_loaded["A1"].value
        wb_loaded.close()
        assert val == "in-memory-test"
    finally:
        if saved_path.exists():
            saved_path.unlink()


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
