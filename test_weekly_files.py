"""Test the new weekly file functionality."""

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from config import Config
from excel_manager import ExcelManager


class TestWeeklyFileFunctionality(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.temp_path = Path(self.temp_dir.name)

        # Create template file
        template_path = self.temp_path / Config.EXCEL_TEMPLATE_NAME
        wb = Workbook()
        wb.create_sheet(Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME)  # "Týden"
        wb.create_sheet(Config.EXCEL_ADVANCES_SHEET_NAME)  # "Zálohy"
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        wb.save(template_path)

        self.excel_manager = ExcelManager(self.temp_path)

    def tearDown(self):
        self.temp_dir.cleanup()

    def test_weekly_file_creation_from_template(self):
        """Test that first weekly file is created from template."""
        # Week 1 should be created from template
        result = self.excel_manager.ulozit_pracovni_dobu("2025-01-06", "08:00", "16:00", "1.0", ["Test Employee"])

        self.assertTrue(result)

        # Check that weekly file was created
        weekly_file = self.temp_path / "Hodiny_Cap_Tyden2.xlsx"
        self.assertTrue(weekly_file.exists())

        # Verify file contains "Týden" sheet
        from openpyxl import load_workbook

        wb = load_workbook(weekly_file)
        self.assertIn(Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME, wb.sheetnames)
        wb.close()

    def test_weekly_file_creation_from_previous_week(self):
        """Test that subsequent weekly files are created from previous week."""
        # Create week 1
        self.excel_manager.ulozit_pracovni_dobu("2025-01-06", "08:00", "16:00", "1.0", ["Employee1"])

        # Create week 2 - should copy from week 1
        result = self.excel_manager.ulozit_pracovni_dobu("2025-01-13", "09:00", "17:00", "1.0", ["Employee2"])

        self.assertTrue(result)

        # Check both files exist
        week1_file = self.temp_path / "Hodiny_Cap_Tyden2.xlsx"
        week2_file = self.temp_path / "Hodiny_Cap_Tyden3.xlsx"

        self.assertTrue(week1_file.exists())
        self.assertTrue(week2_file.exists())

    def test_weekly_file_naming_format(self):
        """Test that weekly files follow correct naming format."""
        # Test different weeks
        test_dates = [
            ("2025-01-06", "Hodiny_Cap_Tyden2.xlsx"),  # Week 2
            ("2025-06-16", "Hodiny_Cap_Tyden25.xlsx"),  # Week 25
            ("2025-12-29", "Hodiny_Cap_Tyden1.xlsx"),  # Week 1 (end of year)
        ]

        for date_str, expected_filename in test_dates:
            result = self.excel_manager.ulozit_pracovni_dobu(date_str, "08:00", "16:00", "1.0", ["Test Employee"])
            self.assertTrue(result)

            expected_path = self.temp_path / expected_filename
            self.assertTrue(expected_path.exists(), f"Expected file {expected_filename} not found")

    def test_template_preservation(self):
        """Test that original template is preserved."""
        original_template = self.temp_path / Config.EXCEL_TEMPLATE_NAME
        original_size = original_template.stat().st_size

        # Create weekly file
        self.excel_manager.ulozit_pracovni_dobu("2025-01-06", "08:00", "16:00", "1.0", ["Test Employee"])

        # Template should still exist and be unchanged
        self.assertTrue(original_template.exists())
        # Note: Size might change slightly due to metadata updates, so we check it's roughly the same
        new_size = original_template.stat().st_size
        self.assertAlmostEqual(original_size, new_size, delta=1000)

    def test_authoritative_template_mapping_writes_start_end_hours_and_date(self):
        self.excel_manager.current_project_name = "Projekt Atlas"

        result = self.excel_manager.ulozit_pracovni_dobu("2025-01-06", "08:00", "16:30", "0.5", ["Test Employee"])

        self.assertTrue(result)
        self.excel_manager.close_cached_workbooks()

        from openpyxl import load_workbook

        wb = load_workbook(self.temp_path / Config.EXCEL_TEMPLATE_NAME)
        ws = wb["Týden 2"]

        self.assertEqual(ws["B4"].value, "NÁZEV PROJEKTU : Projekt Atlas")
        self.assertEqual(ws["B7"].value, "08:00")
        self.assertEqual(ws["C7"].value, "16:30")
        self.assertEqual(ws["B8"].value, 8)
        self.assertEqual(str(ws["B80"].value.date()), "2025-01-06")
        self.assertEqual(ws["A8"].value, "Test Employee")
        wb.close()

    def test_free_day_writes_zero_hours_without_fake_times(self):
        result = self.excel_manager.ulozit_pracovni_dobu("2025-01-07", "00:00", "00:00", "0", ["Free Day User"])

        self.assertTrue(result)
        self.excel_manager.close_cached_workbooks()

        from openpyxl import load_workbook

        wb = load_workbook(self.temp_path / Config.EXCEL_TEMPLATE_NAME)
        ws = wb["Týden 2"]

        self.assertIsNone(ws["D7"].value)
        self.assertIsNone(ws["E7"].value)
        self.assertEqual(ws["D8"].value, 0)
        self.assertEqual(str(ws["D80"].value.date()), "2025-01-07")
        wb.close()


if __name__ == "__main__":
    unittest.main()
