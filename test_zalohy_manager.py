import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from config import Config
from zalohy_manager import ZalohyManager


class TestZalohyManager(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.excel_base_path = Path(self.temp_dir.name)
        self.active_file_path = self.excel_base_path / Config.EXCEL_TEMPLATE_NAME

        wb = Workbook()
        ws = wb.create_sheet(Config.EXCEL_ADVANCES_SHEET_NAME)
        ws["B80"] = Config.DEFAULT_ADVANCE_OPTION_1
        ws["D80"] = Config.DEFAULT_ADVANCE_OPTION_2
        ws["F80"] = Config.DEFAULT_ADVANCE_OPTION_3
        ws["H80"] = Config.DEFAULT_ADVANCE_OPTION_4
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        wb.save(self.active_file_path)

        self.zalohy_manager = ZalohyManager(base_path=str(self.excel_base_path))

    def tearDown(self):
        self.temp_dir.cleanup()

    def _read_cell(self, cell):
        wb = load_workbook(self.active_file_path, data_only=True)
        ws = wb[Config.EXCEL_ADVANCES_SHEET_NAME]
        cell_obj = ws[cell]

        # Handle both single cell and range access
        if isinstance(cell_obj, tuple):
            # If accessing a range, get the first cell
            value = cell_obj[0].value
        else:
            # Single cell access
            value = cell_obj.value

        wb.close()
        return value

    def test_get_option_names(self):
        options = self.zalohy_manager.get_option_names()
        self.assertEqual(len(options), 4)
        self.assertEqual(options[0], Config.DEFAULT_ADVANCE_OPTION_1)

    def test_add_advance_new_employee(self):
        self.zalohy_manager.add_or_update_employee_advance(
            "Nový", 100, "EUR", Config.DEFAULT_ADVANCE_OPTION_1, "2025-01-01"
        )
        self.assertEqual(self._read_cell("A9"), "Nový")
        self.assertEqual(self._read_cell("B9"), 100)
        cell_value = self._read_cell("Z9")
        self.assertIsNotNone(cell_value)
        # Check if cell_value is a datetime object before calling .date()
        if isinstance(cell_value, datetime):
            self.assertEqual(cell_value.date(), datetime(2025, 1, 1).date())
        else:
            # If it's not a datetime, it might be a date object already
            from datetime import date
            if isinstance(cell_value, date):
                self.assertEqual(cell_value, datetime(2025, 1, 1).date())

    def test_add_advances_to_existing_employee(self):
        self.zalohy_manager.add_or_update_employee_advance(
            "Stávající", 50, "EUR", Config.DEFAULT_ADVANCE_OPTION_1, "2025-01-01"
        )
        self.zalohy_manager.add_or_update_employee_advance(
            "Stávající", 25, "EUR", Config.DEFAULT_ADVANCE_OPTION_1, "2025-01-02"
        )
        self.assertEqual(self._read_cell("B9"), 75)

    def test_advance_options_and_currencies(self):
        self.zalohy_manager.add_or_update_employee_advance(
            "Test", 1, "EUR", Config.DEFAULT_ADVANCE_OPTION_1, "2025-01-01"
        )
        self.zalohy_manager.add_or_update_employee_advance(
            "Test", 2, "CZK", Config.DEFAULT_ADVANCE_OPTION_1, "2025-01-01"
        )
        self.zalohy_manager.add_or_update_employee_advance(
            "Test", 3, "EUR", Config.DEFAULT_ADVANCE_OPTION_2, "2025-01-01"
        )
        self.zalohy_manager.add_or_update_employee_advance(
            "Test", 4, "CZK", Config.DEFAULT_ADVANCE_OPTION_2, "2025-01-01"
        )
        self.zalohy_manager.add_or_update_employee_advance(
            "Test", 5, "EUR", Config.DEFAULT_ADVANCE_OPTION_3, "2025-01-01"
        )
        self.zalohy_manager.add_or_update_employee_advance(
            "Test", 6, "CZK", Config.DEFAULT_ADVANCE_OPTION_3, "2025-01-01"
        )
        self.zalohy_manager.add_or_update_employee_advance(
            "Test", 7, "EUR", Config.DEFAULT_ADVANCE_OPTION_4, "2025-01-01"
        )
        self.zalohy_manager.add_or_update_employee_advance(
            "Test", 8, "CZK", Config.DEFAULT_ADVANCE_OPTION_4, "2025-01-01"
        )
        self.assertEqual(self._read_cell("B9"), 1)
        self.assertEqual(self._read_cell("C9"), 2)
        self.assertEqual(self._read_cell("D9"), 3)
        self.assertEqual(self._read_cell("E9"), 4)
        self.assertEqual(self._read_cell("F9"), 5)
        self.assertEqual(self._read_cell("G9"), 6)
        self.assertEqual(self._read_cell("H9"), 7)
        self.assertEqual(self._read_cell("I9"), 8)

    def test_invalid_inputs(self):
        with self.assertRaises(ValueError):
            self.zalohy_manager.add_or_update_employee_advance(
                "Test", -1, "EUR", Config.DEFAULT_ADVANCE_OPTION_1, "2025-01-01"
            )
        with self.assertRaises(ValueError):
            self.zalohy_manager.add_or_update_employee_advance(
                "Test", 100, "USD", Config.DEFAULT_ADVANCE_OPTION_1, "2025-01-01"
            )
        with self.assertRaises(ValueError):
            self.zalohy_manager.add_or_update_employee_advance("Test", 100, "EUR", "Neplatná", "2025-01-01")
        with self.assertRaises(ValueError):
            self.zalohy_manager.add_or_update_employee_advance(
                "Test", 100, "EUR", Config.DEFAULT_ADVANCE_OPTION_1, "neplatne-datum"
            )


if __name__ == "__main__":
    unittest.main()
