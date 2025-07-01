# verify_excel_data.py
import sys
import os
from pathlib import Path
from datetime import datetime, time
import json
import shutil
import re

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# Přidáme kořenový adresář aplikace do sys.path, aby se daly importovat moduly
# Předpokládáme, že tento skript je v kořenovém adresáři HodinyAPP
app_root = Path(__file__).resolve().parent
sys.path.insert(0, str(app_root))

from config import Config
from employee_management import EmployeeManager
from excel_manager import ExcelManager
from zalohy_manager import ZalohyManager
from utils.logger import setup_logger

logger = setup_logger("verify_excel_data")

# --- Konfigurace cest ---
# Cesta k vašemu zdrojovému souboru Excelu
SOURCE_EXCEL_FILE = app_root / "excel" / "Hodiny Cap Týden 3.xlsx"

# Cesty k datům a Excel souborům aplikace (použijeme stejné jako aplikace)
# Tyto cesty jsou definovány v config.py a jsou relativní k BASE_DIR
# Pro účely tohoto skriptu je BASE_DIR = app_root
Config.BASE_DIR = app_root
Config.DATA_PATH = app_root / "data"
Config.EXCEL_BASE_PATH = app_root / "excel"
Config.SETTINGS_FILE_PATH = Config.DATA_PATH / "settings.json"

# Zajistíme, že adresáře existují
Config.DATA_PATH.mkdir(parents=True, exist_ok=True);
Config.EXCEL_BASE_PATH.mkdir(parents=True, exist_ok=True);

# --- Pomocné funkce ---
def load_excel_data(file_path):
    """Načte data z Excel souboru a vrátí je jako slovník listů."""
    data = {}
    try:
        workbook = load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                processed_row = []
                for col_idx, cell_value in enumerate(row):
                    # Zpracování datumu v řádku 80 (index 79)
                    if row_idx == 79 and col_idx in [1, 3, 5, 7, 9, 11, 13]: # Sloupce B, D, F, H, J, L, N
                        if isinstance(cell_value, datetime):
                            processed_row.append(cell_value.strftime("%Y-%m-%d %H:%M:%S")) # Přesný formát pro porovnání
                        elif isinstance(cell_value, str):
                            try:
                                # Zkusíme parsovat různé formáty data
                                if re.match(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", cell_value):
                                    processed_row.append(cell_value)
                                elif re.match(r"^\d{4}-\d{2}-\d{2}$", cell_value):
                                    processed_row.append(f"{cell_value} 00:00:00")
                                elif re.match(r"^\d{1,2}\.\d{1,2}\.\d{4}$", cell_value):
                                    dt_obj = datetime.strptime(cell_value, "%d.%m.%Y")
                                    processed_row.append(dt_obj.strftime("%Y-%m-%d %H:%M:%S"))
                                else:
                                    processed_row.append(cell_value) # Ponecháme, pokud nelze parsovat
                            except ValueError:
                                processed_row.append(cell_value)
                        else:
                            processed_row.append(cell_value)
                    # Zpracování času v řádku 7 (index 6)
                    elif row_idx == 6 and col_idx in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]: # Sloupce B-O
                        if isinstance(cell_value, datetime):
                            processed_row.append(cell_value.strftime("%H:%M:%S")) # Přesný formát pro porovnání
                        elif isinstance(cell_value, time):
                            processed_row.append(cell_value.strftime("%H:%M:%S"))
                        elif isinstance(cell_value, str):
                            try:
                                # Zkusíme parsovat různé formáty času
                                if re.match(r"^\d{2}:\d{2}:\d{2}$", cell_value):
                                    processed_row.append(cell_value)
                                elif re.match(r"^\d{2}:\d{2}$", cell_value):
                                    processed_row.append(f"{cell_value}:00")
                                else:
                                    processed_row.append(cell_value)
                            except ValueError:
                                processed_row.append(cell_value)
                        else:
                            processed_row.append(cell_value)
                    else:
                        processed_row.append(cell_value)
                sheet_data.append(processed_row)
            data[sheet_name] = sheet_data
        workbook.close()
        logger.info(f"Data načtena ze souboru: {file_path}")
        return data
    except FileNotFoundError:
        logger.error(f"Zdrojový soubor Excelu nenalezen: {file_path}")
        return None
    except InvalidFileException:
        logger.error(f"Neplatný soubor Excelu: {file_path}. Zkontrolujte, zda je soubor platný XLSX.")
        return None
    except Exception as e:
        logger.error(f"Chyba při načítání dat ze souboru {file_path}: {e}", exc_info=True)
        return None

def get_cell_value(sheet_data, row, col):
    """Získá hodnotu buňky z načtených dat listu (0-based indexy)."""
    if row < len(sheet_data) and col < len(sheet_data[row]):
        return sheet_data[row][col]
    return None

def format_time_for_app(time_value):
    """Formátuje časovou hodnotu pro použití v aplikaci (HH:MM)."""
    if isinstance(time_value, datetime):
        return time_value.strftime("%H:%M")
    elif isinstance(time_value, time):
        return time_value.strftime("%H:%M")
    elif isinstance(time_value, str):
        # Zkusíme parsovat, pokud je to string, a pak formátovat
        try:
            # Zkusíme formát HH:MM:SS
            dt_obj = datetime.strptime(time_value, "%H:%M:%S")
            return dt_obj.strftime("%H:%M")
        except ValueError:
            try:
                # Zkusíme formát HH:MM
                dt_obj = datetime.strptime(time_value, "%H:%M")
                return dt_obj.strftime("%H:%M")
            except ValueError:
                logger.warning(f"Nelze parsovat čas '{time_value}'. Vrácen prázdný řetězec.")
                return ""
    return ""

def compare_sheets(expected_data, actual_data, sheet_name, key_cells=None):
    """Porovná dva listy Excelu na základě klíčových buněk."""
    logger.info(f"Porovnávám list: {sheet_name}")
    errors = []

    if sheet_name not in actual_data:
        errors.append(f"Chyba: List '{sheet_name}' nebyl nalezen ve výsledném souboru.")
        return errors

    expected_sheet = expected_data.get(sheet_name, [])
    actual_sheet = actual_data.get(sheet_name, [])

    if not expected_sheet and not actual_sheet:
        logger.info(f"List '{sheet_name}' je prázdný v obou souborech, přeskočeno porovnání.")
        return errors

    # Porovnání klíčových buněk, pokud jsou zadány
    if key_cells:
        for r, c, description in key_cells:
            expected_val = get_cell_value(expected_sheet, r, c)
            actual_val = get_cell_value(actual_sheet, r, c)
            if expected_val != actual_val:
                errors.append(f"Rozdíl v listu '{sheet_name}' pro {description} (řádek {r+1}, sloupec {c+1}): Očekáváno '{expected_val}', nalezeno '{actual_val}'.")
            else:
                logger.debug(f"Shoda v listu '{sheet_name}' pro {description}: '{expected_val}'.")
    else:
        # Základní porovnání rozměrů a prvních několika buněk
        if len(expected_sheet) != len(actual_sheet):
            errors.append(f"Rozdíl v počtu řádků v listu '{sheet_name}': Očekáváno {len(expected_sheet)}, nalezeno {len(actual_sheet)}.")
        
        # Porovnání prvních 5 řádků a 10 sloupců pro rychlou kontrolu
        for r_idx in range(min(len(expected_sheet), len(actual_sheet), 5)):
            for c_idx in range(min(len(expected_sheet[r_idx]), len(actual_sheet[r_idx]), 10)):
                expected_val = get_cell_value(expected_sheet, r_idx, c_idx)
                actual_val = get_cell_value(actual_sheet, r_idx, c_idx)
                if str(expected_val).strip() != str(actual_val).strip(): # Porovnáváme jako stringy pro robustnost
                    errors.append(f"Rozdíl v listu '{sheet_name}' v buňce ({r_idx+1}, {c_idx+1}): Očekáváno '{expected_val}', nalezeno '{actual_val}'.")
                    
    return errors

# --- Hlavní logika skriptu ---
def main():
    logger.info("Spouštím ověření dat Excelu pro aplikaci HodinyAPP.")

    # 1. Načtení dat ze zdrojového souboru
    source_data = load_excel_data(SOURCE_EXCEL_FILE)
    if source_data is None:
        logger.error("Nelze pokračovat bez načtení zdrojových dat.")
        return

    # Získání informací o projektu a možnostech záloh z původního souboru
    original_project_name = None
    original_project_start_date = None
    original_project_end_date = None
    original_advance_option1 = Config.DEFAULT_ADVANCE_OPTION_1
    original_advance_option2 = Config.DEFAULT_ADVANCE_OPTION_2

    if Config.EXCEL_ADVANCES_SHEET_NAME in source_data:
        zalohy_sheet_data_source = source_data[Config.EXCEL_ADVANCES_SHEET_NAME]
        original_project_name = get_cell_value(zalohy_sheet_data_source, 78, 0) # A79
        original_project_start_date = get_cell_value(zalohy_sheet_data_source, 80, 2) # C81
        original_project_end_date = get_cell_value(zalohy_sheet_data_source, 80, 3) # D81
        original_advance_option1 = get_cell_value(zalohy_sheet_data_source, 79, 1) or Config.DEFAULT_ADVANCE_OPTION_1 # B80
        original_advance_option2 = get_cell_value(zalohy_sheet_data_source, 79, 3) or Config.DEFAULT_ADVANCE_OPTION_2 # D80

    # 2. Inicializace manažerů aplikace
    # Musíme zajistit, že settings.json existuje a má aktivní soubor nastaven na None
    # aby aplikace vytvořila nový soubor nebo použila šablonu.
    default_settings = Config.get_default_settings()
    default_settings["active_excel_file"] = None # Zajistíme, že se vytvoří nový soubor
    # Nastavíme informace o projektu do výchozího nastavení
    if original_project_name:
        default_settings["project_info"]["name"] = original_project_name
    if original_project_start_date:
        # Převedeme datum na string YYYY-MM-DD, pokud je to datetime objekt
        if isinstance(original_project_start_date, datetime):
            default_settings["project_info"]["start_date"] = original_project_start_date.strftime("%Y-%m-%d")
        else:
            default_settings["project_info"]["start_date"] = str(original_project_start_date)
    if original_project_end_date:
        if isinstance(original_project_end_date, datetime):
            default_settings["project_info"]["end_date"] = original_project_end_date.strftime("%Y-%m-%d")
        else:
            default_settings["project_info"]["end_date"] = str(original_project_end_date)

    try:
        with open(Config.SETTINGS_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump(default_settings, f, indent=4, ensure_ascii=False)
        logger.info("Nastavení aplikace resetováno pro vytvoření nového aktivního souboru.")
    except Exception as e:
        logger.error(f"Chyba při resetování nastavení: {e}", exc_info=True)
        return

    # Inicializujeme Config pro aplikaci (vytvoří adresáře a šablonu, pokud chybí)
    # Toto by mělo zajistit, že active_excel_file bude nastaveno v session
    # a ExcelManager bude mít platnou cestu.
    # Pro účely tohoto skriptu budeme muset simulovat chování before_request
    # nebo přímo inicializovat manažery s aktivním souborem.

    # Znovu načteme nastavení, aby se aktivní soubor vytvořil/nastavil
    # Toto je zjednodušená simulace before_request
    settings = default_settings # Začneme s resetovanými nastaveními
    # Simulace ensure_active_excel_file z app.py
    template_path = Config.EXCEL_BASE_PATH / Config.EXCEL_TEMPLATE_NAME
    if not template_path.exists():
        logger.error(f"Šablona Excelu '{template_path}' neexistuje. Nelze pokračovat.")
        return

    # Vytvoříme nový aktivní soubor, pokud není nastaven
    if not settings.get("active_excel_file"):
        project_name = settings.get("project_info", {}).get("name", "TestProject")
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        new_active_filename = f"{project_name}_{timestamp}.xlsx"
        new_active_file_path = Config.EXCEL_BASE_PATH / new_active_filename
        
        try:
            shutil.copy2(template_path, new_active_file_path)
            settings["active_excel_file"] = new_active_filename
            logger.info(f"Vytvořen nový aktivní soubor pro aplikaci: {new_active_file_path}")
        except Exception as e:
            logger.error(f"Chyba při vytváření nového aktivního souboru: {e}", exc_info=True)
            return

    # Uložíme aktualizované nastavení s novým aktivním souborem
    try:
        with open(Config.SETTINGS_FILE_PATH, "w", encoding="utf-8") as f:
            json.dump(settings, f, indent=4, ensure_ascii=False)
        logger.info("Nastavení aplikace aktualizováno s novým aktivním souborem.")
    except Exception as e:
        logger.error(f"Chyba při ukládání aktualizovaných nastavení: {e}", exc_info=True)
        return

    # Nyní můžeme inicializovat manažery s aktivním souborem
    employee_manager = EmployeeManager(Config.DATA_PATH)
    excel_manager = ExcelManager(Config.EXCEL_BASE_PATH, settings["active_excel_file"], Config.EXCEL_TEMPLATE_NAME)
    zalohy_manager = ZalohyManager(Config.EXCEL_BASE_PATH, settings["active_excel_file"])

    # Nastavíme název projektu v excel_manageru
    if original_project_name:
        excel_manager.set_project_name(original_project_name)

    # Nastavíme možnosti záloh v nově vytvořeném souboru Excelu
    # Musíme otevřít workbook a zapsat do buněk B80 a D80 v listu Zálohy
    wb = None # Inicializace pro finally blok
    try:
        wb = openpyxl.load_workbook(excel_manager.get_active_file_path())
        if Config.EXCEL_ADVANCES_SHEET_NAME not in wb.sheetnames:
            wb.create_sheet(Config.EXCEL_ADVANCES_SHEET_NAME)
        zalohy_sheet = wb[Config.EXCEL_ADVANCES_SHEET_NAME]
        zalohy_sheet["B80"] = original_advance_option1
        zalohy_sheet["D80"] = original_advance_option2
        wb.save(excel_manager.get_active_file_path())
        logger.info(f"Možnosti záloh nastaveny v novém aktivním souboru: {original_advance_option1}, {original_advance_option2}")
    except Exception as e:
        logger.error(f"Chyba při nastavování možností záloh v novém souboru: {e}", exc_info=True)
    finally:
        if wb:
            try:
                wb.close()
            except Exception as close_err:
                logger.warning(f"Chyba při zavírání workbooku po nastavení možností záloh: {close_err}")

    # 3. Simulace zápisu dat do aplikace
    logger.info("Simuluji zápis dat do aplikace...")
    
    # Zpracování listů s pracovní dobou (Týden X)
    for sheet_name, sheet_data in source_data.items():
        if sheet_name.startswith(Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME):
            logger.info(f"Zpracovávám týdenní list: {sheet_name}")
            
            extracted_week_number = None # Inicializace
            week_number_match = re.search(r"Týden (\d+)", sheet_name)
            if week_number_match:
                extracted_week_number = int(week_number_match.group(1))
            else:
                logger.warning(f"Nelze extrahovat číslo týdne z názvu listu '{sheet_name}'. Aplikace si ho vypočítá sama.")

            # Předpokládáme, že řádek 7 obsahuje časy začátku/konce a řádek 80 datumy
            # a řádek 9+ obsahuje jména zaměstnanců a odpracované hodiny
            
            # Získání časů začátku/konce z řádku 7 (0-based index 6)
            # Sloupce B, D, F, H, J, L, N (indexy 1, 3, 5, 7, 9, 11, 13)
            start_times = [get_cell_value(sheet_data, 6, col_idx) for col_idx in [1, 3, 5, 7, 9, 11, 13]]
            end_times = [get_cell_value(sheet_data, 6, col_idx) for col_idx in [2, 4, 6, 8, 10, 12, 14]]

            # Získání datumů z řádku 80 (0-based index 79)
            # Sloupce B, D, F, H, J, L, N (indexy 1, 3, 5, 7, 9, 11, 13)
            dates_str = [get_cell_value(sheet_data, 79, col_idx) for col_idx in [1, 3, 5, 7, 9, 11, 13]]
            
            # Získání jmen zaměstnanců ze sloupce A (index 0) od řádku Config.EXCEL_EMPLOYEE_START_ROW
            employees_in_sheet = []
            for r_idx in range(Config.EXCEL_EMPLOYEE_START_ROW - 1, len(sheet_data)): # -1 pro 0-based index
                employee_name = get_cell_value(sheet_data, r_idx, 0)
                if employee_name and str(employee_name).strip():
                    employees_in_sheet.append(str(employee_name).strip())
                else:
                    break # Předpokládáme, že prázdná buňka znamená konec seznamu zaměstnanců

            # Projdeme každý den v týdnu a každého zaměstnance
            for day_idx in range(7): # 0=Po, 1=Út, ..., 6=Ne
                current_date_obj = dates_str[day_idx]
                logger.debug(f"\n--- Zpracovávám den {day_idx+1} ---")
                logger.debug(f"Původní datum objekt: {current_date_obj}, typ: {type(current_date_obj)}")
                if not current_date_obj:
                    logger.debug("Datum je prázdné, přeskočeno.")
                    continue # Přeskočíme, pokud není datum

                # Převedeme datum na string YYYY-MM-DD
                if isinstance(current_date_obj, datetime):
                    current_date_str = current_date_obj.strftime("%Y-%m-%d")
                elif isinstance(current_date_obj, str):
                    # Zkusíme parsovat různé formáty data a převést na YYYY-MM-DD
                    try:
                        # Zkusíme formát YYYY-MM-DD HH:MM:SS
                        if re.match(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", current_date_obj):
                            dt_obj = datetime.strptime(current_date_obj, "%Y-%m-%d %H:%M:%S")
                            current_date_str = dt_obj.strftime("%Y-%m-%d")
                        # Zkusíme formát YYYY-MM-DD
                        elif re.match(r"^\d{4}-\d{2}-\d{2}$", current_date_obj):
                            current_date_str = current_date_obj
                        # Zkusíme formát DD.MM.YYYY
                        elif re.match(r"^\d{1,2}\.\d{1,2}\.\d{4}$", current_date_obj):
                            dt_obj = datetime.strptime(current_date_obj, "%d.%m.%Y")
                            current_date_str = dt_obj.strftime("%Y-%m-%d")
                        # Zkusíme formát DD/MM/YYYY
                        elif re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", current_date_obj):
                            dt_obj = datetime.strptime(current_date_obj, "%d/%m/%Y")
                            current_date_str = dt_obj.strftime("%Y-%m-%d")
                        else:
                            logger.warning(f"Nelze parsovat datum '{current_date_obj}'. Používám původní řetězec.")
                            current_date_str = str(current_date_obj) # Ponecháme, pokud nelze parsovat
                    except ValueError:
                        logger.warning(f"Chyba při parsování data '{current_date_obj}'. Používám původní řetězec.")
                        current_date_str = str(current_date_obj)
                else:
                    current_date_str = str(current_date_obj) # Může být již string
                logger.debug(f"Datum pro aplikaci: {current_date_str}")

                start_time = start_times[day_idx]
                end_time = end_times[day_idx]
                logger.debug(f"Původní časy: Start={start_time}, End={end_time}")

                start_time_formatted = format_time_for_app(start_time)
                end_time_formatted = format_time_for_app(end_time)
                logger.debug(f"Formátované časy pro aplikaci: Start={start_time_formatted}, End={end_time_formatted}")

                # Pokud jsou časy prázdné, předpokládáme volný den
                is_free_day = not (start_time_formatted and end_time_formatted)

                # Získání odpracovaných hodin pro daný den a zaměstnance
                # Sloupec pro hodiny je base_col_idx + 2 (např. B->D, D->F)
                hours_col_idx = (day_idx * 2) + 1 + 2 # (day_idx * 2) pro posun o 2 sloupce na den, +1 pro B, +2 pro hodiny
                
                for employee_name in employees_in_sheet:
                    # Najdeme řádek zaměstnance v původním listu
                    employee_row_in_source = -1
                    for r_idx in range(Config.EXCEL_EMPLOYEE_START_ROW - 1, len(sheet_data)):
                        if get_cell_value(sheet_data, r_idx, 0) == employee_name:
                            employee_row_in_source = r_idx
                            break
                    
                    if employee_row_in_source != -1:
                        hours_worked_val = get_cell_value(sheet_data, employee_row_in_source, hours_col_idx)
                        
                        try:
                            hours_worked_float = float(hours_worked_val) if hours_worked_val is not None else 0.0
                        except (ValueError, TypeError):
                            hours_worked_float = 0.0 # Treat non-numeric values as 0 hours

                        # Formátované časy pro aplikaci (přesunuto sem)
                        start_time_formatted = format_time_for_app(start_time)
                        end_time_formatted = format_time_for_app(end_time)

                        if is_free_day or hours_worked_float == 0:
                            # Simulujeme volný den
                            logger.debug(f"Simuluji volný den pro {employee_name} dne {current_date_str}")
                            excel_manager.ulozit_pracovni_dobu(
                                current_date_str, "00:00", "00:00", 0.0, [employee_name], week_number=extracted_week_number
                            )
                        elif hours_worked_val is not None:
                            # Simulujeme pracovní dobu
                            logger.debug(f"Simuluji pracovní dobu pro {employee_name} dne {current_date_str}: {start_time_formatted}-{end_time_formatted}")
                            excel_manager.ulozit_pracovni_dobu(
                                current_date_str, start_time_formatted, end_time_formatted, Config.DEFAULT_TIME_CONFIG.lunch_duration, [employee_name], week_number=extracted_week_number
                            )
                    else:
                        logger.warning(f"Zaměstnanec '{employee_name}' nenalezen v původním listu pro záznam hodin.")

    # Zpracování listu Zálohy
    if Config.EXCEL_ADVANCES_SHEET_NAME in source_data:
        logger.info(f"Zpracovávám list: {Config.EXCEL_ADVANCES_SHEET_NAME}")
        zalohy_sheet_data = source_data[Config.EXCEL_ADVANCES_SHEET_NAME]

        # Získání názvů možností záloh z B80 a D80 (0-based indexy 79, 1 a 79, 3)
        option1_name = get_cell_value(zalohy_sheet_data, 79, 1) or Config.DEFAULT_ADVANCE_OPTION_1
        option2_name = get_cell_value(zalohy_sheet_data, 79, 3) or Config.DEFAULT_ADVANCE_OPTION_2

        # Projdeme řádky se zálohami (od Config.EXCEL_EMPLOYEE_START_ROW)
        for r_idx in range(Config.EXCEL_EMPLOYEE_START_ROW - 1, len(zalohy_sheet_data)):
            employee_name = get_cell_value(zalohy_sheet_data, r_idx, 0) # Sloupec A
            if not employee_name or not str(employee_name).strip():
                break # Konec seznamu zaměstnanců

            employee_name = str(employee_name).strip()

            # Sloupce pro zálohy: B (EUR Option1), C (CZK Option1), D (EUR Option2), E (CZK Option2)
            # Datum zálohy je ve sloupci Z (index 25)
            
            # Zpracování Option 1
            amount_eur_opt1 = get_cell_value(zalohy_sheet_data, r_idx, 1)
            amount_czk_opt1 = get_cell_value(zalohy_sheet_data, r_idx, 2)
            
            # Zpracování Option 2
            amount_eur_opt2 = get_cell_value(zalohy_sheet_data, r_idx, 3)
            amount_czk_opt2 = get_cell_value(zalohy_sheet_data, r_idx, 4)

            advance_date_obj = get_cell_value(zalohy_sheet_data, r_idx, 25)
            if isinstance(advance_date_obj, datetime):
                advance_date_str = advance_date_obj.strftime("%Y-%m-%d")
            else:
                advance_date_str = str(advance_date_obj) if advance_date_obj else datetime.now().strftime("%Y-%m-%d") # Použijeme dnešek, pokud chybí

            if amount_eur_opt1:
                logger.debug(f"Simuluji zálohu pro {employee_name}: {amount_eur_opt1} EUR ({option1_name})")
                zalohy_manager.add_or_update_employee_advance(
                    employee_name, float(amount_eur_opt1), "EUR", option1_name, advance_date_str
                )
            if amount_czk_opt1:
                logger.debug(f"Simuluji zálohu pro {employee_name}: {amount_czk_opt1} CZK ({option1_name})")
                zalohy_manager.add_or_update_employee_advance(
                    employee_name, float(amount_czk_opt1), "CZK", option1_name, advance_date_str
                )
            if amount_eur_opt2:
                logger.debug(f"Simuluji zálohu pro {employee_name}: {amount_eur_opt2} EUR ({option2_name})")
                zalohy_manager.add_or_update_employee_advance(
                    employee_name, float(amount_eur_opt2), "EUR", option2_name, advance_date_str
                )
            if amount_czk_opt2:
                logger.debug(f"Simuluji zálohu pro {employee_name}: {amount_czk_opt2} CZK ({option2_name})")
                zalohy_manager.add_or_update_employee_advance(
                    employee_name, float(amount_czk_opt2), "CZK", option2_name, advance_date_str
                )
    else:
        logger.warning(f"List '{Config.EXCEL_ADVANCES_SHEET_NAME}' nebyl nalezen ve zdrojovém souboru.")

    # Ujistíme se, že se všechny změny uloží
    excel_manager.close_cached_workbooks()
    logger.info("Simulace zápisu dat dokončena. Cache workbooků vyčištěna.")

    # 4. Načtení souboru vygenerovaného aplikací
    app_generated_file = excel_manager.get_active_file_path()
    actual_data = load_excel_data(app_generated_file)
    if actual_data is None:
        logger.error("Nelze načíst data z Excel souboru vygenerovaného aplikací.")
        return

    logger.info(f"Názvy listů ve vygenerovaném souboru: {list(actual_data.keys())}")

    # 5. Porovnání dat
    logger.info("Spouštím porovnání dat...")
    all_errors = []

    # Porovnání listů s pracovní dobou
    for sheet_name in source_data.keys():
        if sheet_name.startswith(Config.EXCEL_WEEK_SHEET_TEMPLATE_NAME):
            # Klíčové buňky pro týdenní listy:
            # Řádek 7 (index 6): časy začátku/konce (B, D, F, H, J, L, N)
            # Řádek 80 (index 79): datumy (B, D, F, H, J, L, N)
            # Řádek 79 (index 78): název projektu (B)
            # Sloupce pro hodiny: D, F, H, J, L, N, P (indexy 3, 5, 7, 9, 11, 13, 15)
            
            key_cells_time_sheets = []
            # Times in row 7
            for i, col_idx in enumerate([1, 3, 5, 7, 9, 11, 13]):
                key_cells_time_sheets.append((6, col_idx, f"Start Time Day {i+1}"))
                key_cells_time_sheets.append((6, col_idx + 1, f"End Time Day {i+1}"))
            # Dates in row 80
            for i, col_idx in enumerate([1, 3, 5, 7, 9, 11, 13]):
                key_cells_time_sheets.append((79, col_idx, f"Date Day {i+1}"))
            # Project name
            key_cells_time_sheets.append((78, 1, "Project Name")) # B79

            # Add employee hours for a few employees/days as a sanity check
            # This part would need to be more dynamic for a full comparison
            # For example, check employee "Čáp Jakub" total hours for a specific day
            # This is just an example, a full comparison would iterate through all employees and days
            
            # Example: Check hours for first employee in source data for first day
            first_employee_name = None
            for r_idx in range(Config.EXCEL_EMPLOYEE_START_ROW - 1, len(source_data[sheet_name])):
                emp_name = get_cell_value(source_data[sheet_name], r_idx, 0)
                if emp_name and str(emp_name).strip():
                    first_employee_name = str(emp_name).strip()
                    break
            
            if first_employee_name:
                # Find row for first_employee_name in actual_sheet
                actual_employee_row = -1
                for r_idx in range(Config.EXCEL_EMPLOYEE_START_ROW - 1, len(actual_data.get(sheet_name, []))):
                    if get_cell_value(actual_data[sheet_name], r_idx, 0) == first_employee_name:
                        actual_employee_row = r_idx
                        break
                
                if actual_employee_row != -1:
                    # Check total hours for Day 1 (column D, index 3)
                    key_cells_time_sheets.append((actual_employee_row, 3, f"Hours for {first_employee_name} Day 1"))


            errors = compare_sheets(source_data, actual_data, sheet_name, key_cells=key_cells_time_sheets)
            all_errors.extend(errors)

    # Porovnání listu Zálohy
    if Config.EXCEL_ADVANCES_SHEET_NAME in source_data:
        # Klíčové buňky pro list Zálohy:
        # B80 (index 79, 1) a D80 (index 79, 3) pro názvy možností
        # A79 (index 78, 0) pro název projektu
        # C81 (index 80, 2) pro datum začátku projektu
        # D81 (index 80, 3) pro datum konce projektu
        key_cells_advances_sheet = [
            (79, 1, "Advance Option 1 Name"),
            (79, 3, "Advance Option 2 Name"),
            (78, 0, "Project Name (Advances Sheet)"),
            (80, 2, "Project Start Date (Advances Sheet)"),
            (80, 3, "Project End Date (Advances Sheet)"),
        ]
        
        # Add employee advance amounts for a few employees as a sanity check
        # This would also need to be more dynamic for a full comparison
        # Example: Check advance for first employee in source data for Option 1 EUR
        first_employee_name_adv = None
        for r_idx in range(Config.EXCEL_EMPLOYEE_START_ROW - 1, len(source_data[Config.EXCEL_ADVANCES_SHEET_NAME])):
            emp_name = get_cell_value(source_data[Config.EXCEL_ADVANCES_SHEET_NAME], r_idx, 0);
            if emp_name and str(emp_name).strip():
                first_employee_name_adv = str(emp_name).strip();
                break;
        
        if first_employee_name_adv:
            actual_employee_row_adv = -1;
            for r_idx in range(Config.EXCEL_EMPLOYEE_START_ROW - 1, len(actual_data.get(Config.EXCEL_ADVANCES_SHEET_NAME, []))):
                if get_cell_value(actual_data[Config.EXCEL_ADVANCES_SHEET_NAME], r_idx, 0) == first_employee_name_adv:
                    actual_employee_row_adv = r_idx;
                    break;
            
            if actual_employee_row_adv != -1:
                key_cells_advances_sheet.append((actual_employee_row_adv, 1, f"Advance for {first_employee_name_adv} (EUR Option 1)")); # Column B, index 1
                key_cells_advances_sheet.append((actual_employee_row_adv, 2, f"Advance for {first_employee_name_adv} (CZK Option 1)")); # Column C, index 2
                key_cells_advances_sheet.append((actual_employee_row_adv, 3, f"Advance for {first_employee_name_adv} (EUR Option 2)")); # Column D, index 3
                key_cells_advances_sheet.append((actual_employee_row_adv, 4, f"Advance for {first_employee_name_adv} (CZK Option 2)")); # Column E, index 4
                key_cells_advances_sheet.append((actual_employee_row_adv, 25, f"Advance Date for {first_employee_name_adv}")); # Column Z, index 25


        errors = compare_sheets(source_data, actual_data, Config.EXCEL_ADVANCES_SHEET_NAME, key_cells=key_cells_advances_sheet);
        all_errors.extend(errors);

    if all_errors:
        logger.error("Nalezeny rozdíly mezi zdrojovým a vygenerovaným souborem:");
        for error in all_errors:
            logger.error(f"- {error}");
        print("\nNalezeny rozdíly. Zkontrolujte logy pro detaily.");
    else:
        logger.info("Žádné významné rozdíly nebyly nalezeny. Data se zdají být zapsána správně.");
        print("\nŽádné významné rozdíly nebyly nalezeny. Data se zdají být zapsána správně.");

    # Vyčištění cache workbooků na konci
    excel_manager.close_cached_workbooks();

if __name__ == "__main__":
    main();
