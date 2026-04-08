"""Správa listu záloh (Zálohy) v hlavním Excel souboru.

Funkce:
 - validace vstupů (částka, měna, datum, jméno)
 - kumulativní přičítání částek do správného měnového sloupce dle vybrané "option"
 - zápis data poslední transakce do dedikovaného sloupce
"""
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from config import Config
from services.excel_config_service import get_configured_cells

try:
    from utils.logger import setup_logger

    logger = setup_logger("zalohy_manager")
except ImportError:
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger("zalohy_manager")


class ZalohyManager:
    """Spravuje a aktualizuje list 'Zálohy' v hlavním Excel souboru."""

    def __init__(self, base_path):
        """base_path: adresář s hlavním souborem Hodiny_Cap.xlsx."""
        self.base_path = Path(base_path)
        self.active_filename = Config.EXCEL_TEMPLATE_NAME
        self.active_file_path = self.base_path / self.active_filename
        self.zalohy_sheet_name = Config.EXCEL_ADVANCES_SHEET_NAME
        self.employee_start_row = Config.EXCEL_EMPLOYEE_START_ROW
        self.valid_currencies = ["EUR", "CZK"]
        self.date_column_index = 26
        logger.info("ZalohyManager inicializován pro soubor: %s", self.active_file_path)

    def _load_dynamic_config(self):
        """Načte dynamickou konfiguraci pro ukládání do XLSX souborů."""
        from services.excel_config_service import load_dynamic_excel_config

        return load_dynamic_excel_config()

    def _get_cell_coordinates(self, field_key, sheet_name=None):
        """Vrátí seznam (row, col) souřadnic pro daný field z dynamické konfigurace.

        Args:
            field_key: Klíč pole z konfigurace (např. 'employee_name', 'amount_eur')
            sheet_name: Název listu, pokud chceme ověřit shodu

        Returns:
            list: Seznam (row, col) souřadnic nebo prázdný seznam pokud není nakonfigurováno
        """
        return get_configured_cells("advances", field_key, self.active_filename, sheet_name=sheet_name)

    def _get_active_workbook(self, read_only=False):
        if not self.active_file_path.exists():
            raise FileNotFoundError(f"Soubor '{self.active_filename}' neexistuje.")
        try:
            return load_workbook(filename=self.active_file_path, read_only=read_only, data_only=True)
        except Exception as e:
            raise IOError(f"Nepodařilo se otevřít soubor '{self.active_filename}'.") from e

    def _save_workbook(self, workbook):
        try:
            workbook.save(self.active_file_path)
        except Exception as e:
            raise IOError(f"Nepodařilo se uložit změny do souboru '{self.active_filename}'.") from e

    def _validate_inputs(self, employee_name, amount, currency, date_str):
        if not isinstance(amount, (int, float)) or amount <= 0:
            raise ValueError("Částka musí být kladné číslo.")
        if currency not in self.valid_currencies:
            raise ValueError("Neplatná měna.")
        if not isinstance(employee_name, str) or not employee_name.strip():
            raise ValueError("Jméno zaměstnance nemůže být prázdné.")
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError as e:
            raise ValueError("Neplatný formát data.") from e

    @staticmethod
    def _resolve_option_coordinate(option_coords, option_index):
        if option_index < len(option_coords):
            return option_coords[option_index]
        return None

    def add_or_update_employee_advance(self, employee_name, amount, currency, option, date):
        """Přičte zálohu zaměstnanci (vytvoří řádek pokud chybí)."""
        workbook = None
        try:
            self._validate_inputs(employee_name, amount, currency, date)
            workbook = self._get_active_workbook(read_only=False)
            sheet = workbook[self.zalohy_sheet_name]
            options = self.get_option_names()
            if option not in options:
                raise ValueError(f"Neplatná volba zálohy: {option}")

            row = self._get_or_create_employee_row(sheet, employee_name)
            option_index = options.index(option)

            # Zkus použít dynamickou konfiguraci pro částky
            amount_field = f"amount_{currency.lower()}"
            amount_coords = self._get_cell_coordinates(amount_field, self.zalohy_sheet_name)

            if amount_coords:
                target_coordinate = self._resolve_option_coordinate(amount_coords, option_index)
                if target_coordinate:
                    amount_row, amount_col = target_coordinate
                    actual_row = row if amount_row <= self.employee_start_row else amount_row
                    self._update_advance_cell(sheet, actual_row, amount_col, amount)
                    logger.info(
                        "Částka %s %s zapsána do buňky %s%d (dynamická konfigurace)",
                        amount,
                        currency,
                        chr(64 + amount_col),
                        actual_row,
                    )
                else:
                    raise ValueError(f"Chybí mapping pro volbu zálohy {option} a měnu {currency}.")
            else:
                # Fallback na původní logiku
                column_index = 2 + (option_index * 2) + (1 if currency == "CZK" else 0)
                self._update_advance_cell(sheet, row, column_index, amount)
                logger.info(
                    "Částka %s %s zapsána na řádek %d, sloupec %d (původní logika)", amount, currency, row, column_index
                )

            self._update_date_cell(sheet, row, date)

            self._save_workbook(workbook)
            return True
        except (FileNotFoundError, ValueError, IOError) as e:
            logger.error("Chyba při ukládání zálohy: %s", e, exc_info=True)
            raise
        finally:
            if workbook:
                workbook.close()

    def _get_or_create_employee_row(self, sheet, employee_name):
        # Zkus použít dynamickou konfiguraci pro pozici jména zaměstnance
        employee_name_coords = self._get_cell_coordinates("employee_name", self.zalohy_sheet_name)

        if employee_name_coords:
            # Pokud je nakonfigurováno, použij konfigurovanou pozici
            config_row, config_col = employee_name_coords[0]  # Použij první lokaci

            # Hledej existující zaměstnance od konfigurované pozice
            for row in range(config_row, sheet.max_row + 2):
                cell = sheet.cell(row=row, column=config_col)
                if cell.value == employee_name:
                    return row
                if cell.value is None:
                    cell.value = employee_name
                    logger.info("Zaměstnanec %s přidán na řádek %d (dynamická konfigurace)", employee_name, row)
                    return row
        else:
            # Fallback na původní logiku
            for row in range(self.employee_start_row, sheet.max_row + 2):
                cell = sheet.cell(row=row, column=1)
                if cell.value == employee_name:
                    return row
                if cell.value is None:
                    cell.value = employee_name
                    logger.info("Zaměstnanec %s přidán na řádek %d (původní logika)", employee_name, row)
                    return row

        return sheet.max_row + 1  # Should not be reached in practice

    def _update_advance_cell(self, sheet, row, column, amount):
        target_cell = sheet.cell(row=row, column=column)
        if isinstance(target_cell, MergedCell):
            return

        current_value = 0.0
        try:
            if target_cell.value is not None:
                current_value = float(target_cell.value)
        except (ValueError, TypeError):
            current_value = 0.0

        target_cell.value = current_value + amount
        target_cell.number_format = "#,##0.00"

    def _update_date_cell(self, sheet, row, date_str):
        # Zkus použít dynamickou konfiguraci pro datum
        date_coords = self._get_cell_coordinates("date", self.zalohy_sheet_name)

        if date_coords:
            date_row, date_col = date_coords[0]
            actual_row = row if date_row <= self.employee_start_row else date_row
            date_cell = sheet.cell(row=actual_row, column=date_col)
            if not isinstance(date_cell, MergedCell):
                date_cell.value = datetime.strptime(date_str, "%Y-%m-%d").date()
                date_cell.number_format = "DD.MM.YYYY"
                logger.info("Datum zapsáno do buňky %s%d (dynamická konfigurace)", chr(64 + date_col), actual_row)
        else:
            # Fallback na původní logiku
            date_cell = sheet.cell(row=row, column=self.date_column_index)
            if not isinstance(date_cell, MergedCell):
                date_cell.value = datetime.strptime(date_str, "%Y-%m-%d").date()
                date_cell.number_format = "DD.MM.YYYY"
                logger.info("Datum zapsáno na řádek %d, sloupec %d (původní logika)", row, self.date_column_index)

    def get_option_names(self):
        """Názvy 4 možností čtené z buněk (fallback na default)."""
        default_options = (
            Config.DEFAULT_ADVANCE_OPTION_1,
            Config.DEFAULT_ADVANCE_OPTION_2,
            Config.DEFAULT_ADVANCE_OPTION_3,
            Config.DEFAULT_ADVANCE_OPTION_4,
        )
        workbook = None
        try:
            workbook = self._get_active_workbook(read_only=True)
            if self.zalohy_sheet_name in workbook.sheetnames:
                sheet = workbook[self.zalohy_sheet_name]
                option_type_coords = self._get_cell_coordinates("option_type", self.zalohy_sheet_name)
                if option_type_coords:
                    options = []
                    for index, default in enumerate(default_options):
                        option_coordinate = self._resolve_option_coordinate(option_type_coords, index)
                        if not option_coordinate:
                            options.append(default)
                            continue
                        option_row, option_col = option_coordinate
                        options.append(str(sheet.cell(row=option_row, column=option_col).value or default).strip())
                    return tuple(options)
                return tuple(
                    str(sheet[cell].value or default).strip()
                    for cell, default in zip(["B80", "D80", "F80", "H80"], default_options)
                )
            return default_options
        except (FileNotFoundError, IOError) as e:
            logger.error("Chyba při čtení názvů možností: %s", e, exc_info=True)
            return default_options
        finally:
            if workbook:
                workbook.close()
